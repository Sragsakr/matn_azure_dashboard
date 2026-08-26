"""
Delivery Manager - Streamlit Web Dashboard (entry point)
----------------------------------------------------------
Pulls live data from Azure DevOps ("Hoteliana") and renders an interactive
Delivery Manager dashboard in the browser. Refresh pulls fresh Azure data.

If the AZDO_PAT environment variable is not set, it falls back to whatever is
already in the local workbook's "Raw Data" sheet so the UI still opens without
secrets. Set AZDO_PAT (env var / secret) to enable live pulls.

Phase 2 (multipage restructure): this file is the native Streamlit
multipage entry point. It owns global page config, theme/language
initialization, Azure/workbook data loading, and shared precomputed
analysis — then hands off rendering to the page scripts under pages/ via
st.navigation(). A streamlit-option-menu sidebar drives page selection
(grouped exactly like the previous st.sidebar.radio groups) while
st.navigation's own widget stays hidden. All rendering logic, CSS
injection, and Altair chart theming are unchanged from Phase 1 — this is
a structural move only.
"""

import os
import base64
from urllib.parse import quote

import openpyxl
import requests
import streamlit as st
from streamlit_option_menu import option_menu

import dashboard_theme
from dashboard_styles import apply_theme
from core.analysis import (
    ORG, PROJECT, DELIVERY_TYPES, DEV_TYPES, STALE_DAYS, PB,
    _leaf, _sprint, _parse_date, _category, _wi,
    is_open, is_active,
    item_metrics, type_progress, scope_metrics, percent, ribbon,
    weekly_creation_closure, sprint_summary_df as _sprint_summary_df,
    team_df as _team_df, area_df as _area_df,
)
from core.ui_helpers import (
    tr as _ui_tr,
    column_label as _ui_column_label,
    localized_frame as _ui_localized_frame,
    localized_label as _ui_localized_label,
    percentage_columns as _ui_percentage_columns,
    delivery_action as _ui_delivery_action,
    theme_chart as _ui_theme_chart,
    kpi_card,
    pat as _pat,
)
from components.filter_bar import render_filter_bar

# ---------------------------------------------------------------- constants
API_VERSION = "7.1"
WORKBOOK = "Delivery_Manager_Dashboard.xlsx"


# ---------------------------------------------------------------- data layer
def _auth_header(pat):
    return {
        "Authorization": "Basic " + base64.b64encode(f":{pat}".encode()).decode(),
        "Content-Type": "application/json",
    }


def _state_categories(base, headers, work_item_types):
    """Return Azure state category keyed by (work-item type, state name)."""
    categories = {}
    for work_type in work_item_types:
        try:
            response = requests.get(
                f"{base}/wit/workitemtypes/{quote(work_type, safe='')}/states"
                f"?api-version={API_VERSION}",
                headers=headers,
                timeout=15,
            )
            response.raise_for_status()
            for state in response.json().get("value", []):
                categories[(work_type, state.get("name"))] = state.get("category")
        except requests.RequestException:
            # A restricted PAT may read work items but not process metadata.
            # Classification then falls back to the known state names below.
            continue
    return categories


def pull_from_azure():
    """Pull and enrich the latest Azure DevOps work items."""
    pat = _pat()
    if not pat:
        raise ValueError("AZDO_PAT not set — cannot pull live data")

    base = f"https://dev.azure.com/{ORG}/{PROJECT}/_apis"
    hdr = _auth_header(pat)
    query = {
        "query": (
            "SELECT [System.Id] FROM WorkItems "
            f"WHERE [System.TeamProject] = '{PROJECT}' "
            "AND [System.WorkItemType] <> '' "
            "ORDER BY [System.ChangedDate] DESC"
        )
    }
    resp = requests.post(f"{base}/wit/wiql?api-version={API_VERSION}",
                         headers=hdr, json=query, timeout=30)
    resp.raise_for_status()
    ids = [item["id"] for item in resp.json().get("workItems", [])]

    fields = [
        "System.Id", "System.Title", "System.WorkItemType", "System.State",
        "System.AssignedTo", "System.IterationPath", "System.AreaPath",
        "System.Parent", "Microsoft.VSTS.Scheduling.StoryPoints",
        "Microsoft.VSTS.Common.Priority", "System.CreatedDate",
        "System.ChangedDate", "System.Tags", "System.BoardColumn",
        "System.BoardColumnDone", "System.BoardLane",
    ]
    all_items = []
    for i in range(0, len(ids), 200):
        chunk = ids[i:i + 200]
        resp = requests.post(
            f"{base}/wit/workitemsbatch?api-version={API_VERSION}",
            headers=hdr, json={"ids": chunk, "fields": fields}, timeout=30)
        resp.raise_for_status()
        all_items.extend(resp.json().get("value", []))

    work_types = {
        item.get("fields", {}).get("System.WorkItemType")
        for item in all_items
        if item.get("fields", {}).get("System.WorkItemType")
    }
    categories = _state_categories(base, hdr, work_types)
    for item in all_items:
        item_fields = item.get("fields", {})
        item["_state_category"] = categories.get((
            item_fields.get("System.WorkItemType"),
            item_fields.get("System.State"),
        ))
    return all_items


def read_workbook_items():
    """Fallback: read normalized items from the local workbook's Raw Data.
    Returns [] when the workbook is absent (e.g. on a cloud host)."""
    if not os.path.exists(WORKBOOK):
        return []
    wb = openpyxl.load_workbook(WORKBOOK, data_only=True)
    ws = wb["Raw Data"]
    header = [c.value for c in ws[5]]
    idx = {name: i for i, name in enumerate(header) if name}
    items = []
    for row in ws.iter_rows(min_row=6, values_only=True):
        if not row or row[0] is None and (row[1] is None):
            continue
        def g(name, default=None):
            i = idx.get(name)
            return row[i] if i is not None and i < len(row) else default
        created = _parse_date(g("Created Date"))
        state = g("State") or "Unknown"
        items.append({
            "id": g("Work Item ID"),
            "title": g("Title") or "",
            "type": g("Work Item Type") or "Unknown",
            "state": state,
            "state_category": _category(state, g("State Category")),
            "board_column": g("Board Column") or state,
            "board_column_done": bool(g("Board Column Done")),
            "board_lane": g("Board Lane") or "Default",
            "assignee": g("Assigned To") or "Unassigned",
            "sprint": _sprint(g("Iteration Path")),
            "area": _leaf(g("Area Path") or PROJECT),
            "sp": g("Story Points"),
            "priority": g("Priority"),
            "created": created,
            "changed": _parse_date(g("Changed Date")),
            "parent": g("Parent ID"),
            "tags": [t.strip() for t in str(g("Tags") or "").split(";") if t.strip()],
            "url": f"https://dev.azure.com/{ORG}/{PROJECT}/_workitems/edit/{g('Work Item ID')}",
        })
    return items


def load_items(force_pull=True):
    """Try live Azure pull; fall back to workbook Raw Data.
    Returns (items, mode). mode in {'live','workbook','empty'}."""
    if force_pull and _pat():
        try:
            raw = pull_from_azure()
            items = [_wi(w) for w in raw]
            return items, "live"
        except Exception as exc:
            # keep the real error visible (for debugging), fall back to cache
            st.session_state["last_pull_error"] = str(exc)
    items = read_workbook_items()
    if items:
        return items, "workbook"
    return [], "empty"


# ============================================================ UI HELPERS
# Thin is_ar-bound wrappers around core.ui_helpers, kept so the rest of this
# file can call these with the original zero-arg-for-locale call sites.
# Pages must import the underlying core.ui_helpers functions directly with
# an explicit is_ar argument instead — see core/ui_helpers.py docstring for
# why importing these wrappers from app would break page navigation.
def tr(english, arabic):
    return _ui_tr(english, arabic, is_ar)


def column_label(name):
    return _ui_column_label(name, is_ar)


def localized_frame(frame):
    return _ui_localized_frame(frame, is_ar)


def localized_label(name):
    return _ui_localized_label(name, is_ar)


def percentage_columns(*names):
    return _ui_percentage_columns(*names, is_ar=is_ar)


def delivery_action(items, unassigned):
    return _ui_delivery_action(items, unassigned, is_ar)


def theme_chart(chart):
    return _ui_theme_chart(chart)


# ---------------------------------------------------------------- rendering
st.set_page_config(
    page_title="Delivery Manager — Hoteliana",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
)

is_ar = st.session_state.get("language_selector", "العربية") == "العربية"
theme_mode = "dark" if st.session_state.get("theme_selector", "Dark") == "Dark" else "light"
apply_theme(theme_mode, is_ar)
chart_theme = dashboard_theme.chart_theme(theme_mode)
ACCENT = dashboard_theme.ACCENTS

st.sidebar.markdown(
    """
    <div class="sidebar-brand">
      <div class="brand-mark">M</div>
      <div><strong>MATN DELIVERY</strong><span>Hoteliana workspace</span></div>
    </div>
    """,
    unsafe_allow_html=True,
)
st.sidebar.segmented_control(
    tr("Language", "اللغة"),
    ["العربية", "English"],
    key="language_selector",
    default="العربية",
    label_visibility="collapsed",
)
st.sidebar.segmented_control(
    tr("Theme", "المظهر"),
    ["Dark", "Light"],
    key="theme_selector",
    default="Dark",
    label_visibility="collapsed",
)
st.sidebar.markdown(
    f'<div class="sidebar-label">{tr("DATA CONTROL", "التحكم بالبيانات")}</div>',
    unsafe_allow_html=True,
)

st.markdown(
    f"""
    <section class="enterprise-header">
      <div>
        <div class="header-kicker">MATN SOLUTIONS</div>
        <h1>{tr("Strategic Delivery Command Center", "مركز القيادة الاستراتيجي للتسليم")}</h1>
      </div>
      <div class="azure-pill"><span class="azure-dot"></span>Azure DevOps · {PROJECT}</div>
    </section>
    """,
    unsafe_allow_html=True,
)

user_missing = not _pat()
if user_missing:
    st.sidebar.warning(tr(
        "AZDO_PAT is not set — showing the latest workbook snapshot.",
        "رمز AZDO_PAT غير مضاف — يتم عرض آخر نسخة محفوظة من البيانات.",
    ))

refresh = st.sidebar.button(
    tr("↻  Sync Azure DevOps", "↻  مزامنة Azure DevOps"),
    type="primary",
)

if refresh:
    st.session_state.pop("azure_items", None)
    st.session_state.pop("azure_data_mode", None)

if "azure_items" not in st.session_state:
    loading_text = tr(
        "Syncing Azure DevOps data...",
        "جارٍ مزامنة بيانات Azure DevOps...",
    ) if _pat() else tr(
        "Loading saved data...",
        "جارٍ تحميل البيانات المحفوظة...",
    )
    with st.spinner(loading_text):
        loaded_items, loaded_mode = load_items(force_pull=bool(_pat()))
    st.session_state["azure_items"] = loaded_items
    st.session_state["azure_data_mode"] = loaded_mode

items = st.session_state["azure_items"]
data_mode = st.session_state["azure_data_mode"]

# Phase 6: a pull failure used to surface only as a small st.sidebar.error,
# easy to miss since the sidebar isn't always in view (especially once
# collapsed on narrow screens). Mirror the same failure as a full-width
# banner in the main content area, directly under the header, so it's
# visible regardless of sidebar state — on every rerun while the failure
# is current, not just immediately after the user clicks Refresh.
if data_mode != "live" and st.session_state.get("last_pull_error"):
    st.error(
        tr(
            "⚠ Azure DevOps sync failed — showing the last saved snapshot instead.",
            "⚠ فشلت مزامنة Azure DevOps — يتم عرض آخر نسخة محفوظة بدلاً من ذلك.",
        )
        + f"\n\n`{st.session_state['last_pull_error']}`\n\n"
        + tr(
            "**What to try:** confirm `AZDO_PAT` is set and has not expired, then press "
            "**↻ Sync Azure DevOps** in the sidebar to retry.",
            "**للمحاولة:** تأكد من ضبط `AZDO_PAT` وأنه لم تنتهِ صلاحيته، ثم اضغط "
            "**↻ مزامنة Azure DevOps** في الشريط الجانبي لإعادة المحاولة.",
        ),
        icon="🚨",
    )

if refresh:
    if data_mode == "live":
        st.sidebar.success(tr(
            f"Synced {len(items)} items from Azure.",
            f"تمت مزامنة {len(items)} عنصر من Azure.",
        ))
    else:
        error = st.session_state.get("last_pull_error", "Azure credentials are unavailable.")
        st.sidebar.error(tr(
            f"Sync failed; saved data is shown. {error}",
            f"فشلت المزامنة؛ يتم عرض البيانات المحفوظة. {error}",
        ))

dev = [i for i in items if i["type"] in DEV_TYPES]

if not dev:
    st.sidebar.warning("No data loaded.")
    st.error("No dev work items found.")
    missing_pat = not _pat()
    if missing_pat:
        st.info("`AZDO_PAT` secret is not set. Set it in the hosting platform's secrets "
                "(Streamlit Cloud → Settings → Secrets) or as the GitHub Actions secret, "
                "then press **Refresh from Azure DevOps**.")
    # Show the underlying pull failure so it isn't a silent "nothing on screen"
    if st.session_state.get("last_pull_error"):
        st.warning(f"Last Azure pull failed: {st.session_state['last_pull_error']}")
    st.info("If you are self-hosting without a workbook, the app needs live Azure access "
            "via AZDO_PAT to load any data.")
    st.stop()

# Phase 4: persistent global filter bar — rendered once here, below the
# enterprise header and above every page's content. Filters `dev` before
# all_m/scope/verdict/prog are derived from it, so the filtered result
# flows into app_ctx and every page automatically sees filtered data
# through the same app_ctx mechanism, with zero per-page changes.
dev = render_filter_bar(dev, is_ar=is_ar)

if not dev:
    st.info(tr(
        "No items match the current filters. Adjust or clear them above.",
        "لا توجد عناصر مطابقة للفلاتر الحالية. عدّلها أو امسحها أعلاه.",
    ))

# Precompute shared analysis once (executive-only metrics used by sidebar).
all_m = item_metrics(dev)
scope = scope_metrics([i for i in dev if i["sprint"] != PB])
verdict, color = ribbon(scope["scope_pct"], scope["task_pct"], all_m["unassigned"], all_m["stale"])
prog = type_progress(dev)

# Stash everything page scripts need — Streamlit's official pattern for
# sharing entrypoint-computed data with pages/ scripts is session_state
# (each page runs as its own script; only widgets with `key=` and explicit
# session_state entries are guaranteed visible across that boundary).
st.session_state["app_ctx"] = {
    "is_ar": is_ar,
    "theme_mode": theme_mode,
    "chart_theme": chart_theme,
    "ACCENT": ACCENT,
    "items": items,
    "data_mode": data_mode,
    "dev": dev,
    "all_m": all_m,
    "scope": scope,
    "verdict": verdict,
    "color": color,
    "prog": prog,
}

# ---- Sidebar navigation: one section per tab, mirroring the Excel workbook
PAGES = {
    "Executive Dashboard": ("◈  Executive overview", "◈  النظرة التنفيذية", "pages/1_executive_dashboard.py"),
    "Sprint Summary": ("⏱  Sprint summary", "⏱  ملخص السبرينت", "pages/2_sprint_summary.py"),
    "Sprint Board": ("▦  Sprint board", "▦  لوحة السبرينت", "pages/3_sprint_board.py"),
    "Tag Analysis": ("#  Tag analysis", "#  تحليل الوسوم", "pages/4_tag_analysis.py"),
    "Team Analysis": ("👥  Team delivery", "👥  أداء الفريق", "pages/5_team_analysis.py"),
    "Area Analysis": ("◇  Area analysis", "◇  تحليل المجالات", "pages/6_area_analysis.py"),
    "Active Now": ("⚡  Active now", "⚡  العمل الحالي", "pages/7_active_now.py"),
    "Risks & Aging": ("⚠  Risks & aging", "⚠  المخاطر والتقادم", "pages/8_risks_and_aging.py"),
    "Data Quality": ("✓  Data quality", "✓  جودة البيانات", "pages/9_data_quality.py"),
    "Repository Intelligence": ("📊  Repository intelligence", "📊  ذكاء المستودعات", "pages/10_repository_intelligence.py"),
    "Releases": ("🏆  Releases", "🏆  الإصدارات", "pages/11_releases.py"),
    "Raw Data": ("📁  Raw data", "📁  البيانات الخام", "pages/12_raw_data.py"),
}
# Grouped purely for sidebar presentation — every PAGES key appears in exactly
# one group. The streamlit-option-menu sidebar renders these groups (each as
# its own labeled option_menu, mirroring the previous st.sidebar.radio
# groups) and writes the shared "dashboard_page" key that st.navigation below
# dispatches on.
PAGE_GROUPS = [
    (("Overview", "نظرة عامة"), ["Executive Dashboard", "Sprint Summary", "Sprint Board"]),
    (("Delivery & Team", "التسليم والفريق"), ["Tag Analysis", "Team Analysis", "Area Analysis", "Active Now"]),
    (("Quality & Risk", "الجودة والمخاطر"), ["Risks & Aging", "Data Quality"]),
    (("Data & Repositories", "البيانات والمستودعات"), ["Repository Intelligence", "Releases", "Raw Data"]),
]
assert sorted(sum((keys for _, keys in PAGE_GROUPS), [])) == sorted(PAGES), \
    "PAGE_GROUPS must partition PAGES exactly"

if "dashboard_page" not in st.session_state:
    st.session_state["dashboard_page"] = next(iter(PAGES))

def _sync_dashboard_page_from_menu(group_option_key, group_keys, labels):
    selection = st.session_state[group_option_key]
    st.session_state["dashboard_page"] = group_keys[labels.index(selection)]


for (label_en, label_ar), group_keys in PAGE_GROUPS:
    st.sidebar.markdown(
        f'<div class="sidebar-label">{tr(label_en, label_ar)}</div>',
        unsafe_allow_html=True,
    )
    group_option_key = f"nav_group_{label_en}"
    labels = [PAGES[key][1 if is_ar else 0] for key in group_keys]
    current = st.session_state["dashboard_page"]
    # streamlit-option-menu is a custom component: once mounted under a given
    # `key`, its own remembered selection persists across reruns regardless
    # of default_index, EXCEPT when manual_select forces it — so force the
    # index only for the group that currently owns the active page (keeping
    # every group pinned to its own last choice, exactly like the previous
    # st.sidebar.radio groups). on_change (not the return value) is what
    # updates "dashboard_page", so a stale/defaulted return value from an
    # inactive group never overwrites the real selection.
    force_index = group_keys.index(current) if current in group_keys else 0
    option_menu(
        menu_title=None,
        options=labels,
        default_index=force_index,
        manual_select=force_index if current in group_keys else None,
        key=group_option_key,
        on_change=lambda key, gk=group_keys, lb=labels: _sync_dashboard_page_from_menu(key, gk, lb),
        styles={
            "container": {"padding": "0", "background-color": "transparent"},
            "icon": {"display": "none"},
            "nav-link": {
                "font-size": "0.74rem", "padding": "0.4rem 0.5rem",
                "margin": "0 0 1px 0", "border-radius": "5px",
            },
            "nav-link-selected": {"background-color": "var(--surface2)", "font-weight": "700"},
        },
    )

page = st.session_state["dashboard_page"]
st.sidebar.markdown(
    f'<div class="sidebar-label">{tr("SCOPE SNAPSHOT", "ملخص النطاق")}</div>',
    unsafe_allow_html=True,
)
source_col, scope_col = st.sidebar.columns(2)
source_col.metric(tr("Source", "المصدر"), tr(data_mode.title(), "مباشر" if data_mode == "live" else "محفوظ"))
scope_col.metric(tr("Delivery", "التسليم"), len(dev))
st.sidebar.caption(tr(
    f"{len(items):,} total Azure work items",
    f"إجمالي عناصر Azure: {len(items):,}",
))

st.caption(tr(
    f"Live synchronization with {ORG}/{PROJECT}." if data_mode == "live" else "Showing the latest saved Azure snapshot.",
    f"مزامنة مباشرة مع {ORG}/{PROJECT}." if data_mode == "live" else "يتم عرض آخر نسخة محفوظة من بيانات Azure.",
))
st.divider()

# ---- native multipage routing: the option_menu sidebar above already picked
# `page`; st.navigation's own nav widget is hidden and only used to register
# pages/run the selected one, keeping URLs/page identity native to Streamlit.
_path_by_key = {title_key: path for title_key, (_, _, path) in PAGES.items()}
_pages = [st.Page(path, title=title_key) for title_key, path in _path_by_key.items()]
pg = st.navigation(_pages, position="hidden")

# st.navigation() resolves to whichever page matches the current browser URL
# (the first page on first load). _active_page tracks the key we last routed
# to via our own sidebar, so we only call st.switch_page on a real selection
# change — st.switch_page triggers an immediate rerun, so an unconditional
# call here would loop.
if st.session_state.get("_active_page") != page:
    st.session_state["_active_page"] = page
    st.switch_page(_path_by_key[page])
pg.run()
