"""
Delivery Manager - Streamlit Web Dashboard
------------------------------------------
Pulls live data from Azure DevOps ("Hoteliana") and renders an interactive
Delivery Manager dashboard in the browser. Refresh pulls fresh Azure data.

If the AZDO_PAT environment variable is not set, it falls back to whatever is
already in the local workbook's "Raw Data" sheet so the UI still opens without
secrets. Set AZDO_PAT (env var / secret) to enable live pulls.
"""

import os
import base64
import datetime as dt
from collections import Counter, defaultdict
from urllib.parse import quote

import openpyxl
import pandas as pd
import requests
import altair as alt
import streamlit as st

import dashboard_theme
from azure_repo_activity import AzureRepoActivityClient, contributor_rows
from dashboard_styles import apply_theme, section_header

# ---------------------------------------------------------------- constants
ORG = "matnsolutions"
PROJECT = "Hoteliana"
API_VERSION = "7.1"
WORKBOOK = "Delivery_Manager_Dashboard.xlsx"

# Fallbacks keep older workbook caches useful. Live pulls use Azure's canonical
# state categories, so newly-added custom states are classified correctly.
DONE = {"Closed", "Done", "Resolved", "Completed"}
ACTIVE = {"Active", "In Progress", "Committed"}
TERMINAL_CATEGORIES = {"Completed", "Removed"}
DELIVERY_TYPES = ("Epic", "Feature", "User Story", "Task", "Bug")
DEV_TYPES = set(DELIVERY_TYPES)
STALE_DAYS = 14
PB = "Product Backlog"  # project-only iteration label


# ---------------------------------------------------------------- data layer
def _leaf(path):
    if not path:
        return ""
    return str(path).replace("\\\\", "\\").split("\\")[-1]


def _sprint(path):
    if not path:
        return PB
    p = str(path).replace("\\\\", "\\")
    return p.split("\\")[-1] if "\\" in p else PB


def _parse_date(v):
    if not v:
        return None
    if isinstance(v, dt.datetime):
        return v.date()
    if isinstance(v, dt.date):
        return v
    try:
        return dt.datetime.fromisoformat(str(v)[:10]).date()
    except ValueError:
        return None


def _auth_header(pat):
    return {
        "Authorization": "Basic " + base64.b64encode(f":{pat}".encode()).decode(),
        "Content-Type": "application/json",
    }


def _pat():
    """PAT from Streamlit secrets first, then env var."""
    try:
        s = st.secrets.get("AZDO_PAT")
        if s:
            return s
    except Exception:
        pass
    return os.environ.get("AZDO_PAT")


def _state_categories(base, headers, work_item_types):
    """Return Azure state category keyed by (work-item type, state name)."""
    categories = {}
    for work_type in work_item_types:
        try:
            response = requests.get(
                f"{base}/wit/workitemtypes/{quote(work_type, safe='')}/states"
                f"?api-version={API_VERSION}",
                headers=headers,
                timeout=15,
            )
            response.raise_for_status()
            for state in response.json().get("value", []):
                categories[(work_type, state.get("name"))] = state.get("category")
        except requests.RequestException:
            # A restricted PAT may read work items but not process metadata.
            # Classification then falls back to the known state names below.
            continue
    return categories


def pull_from_azure():
    """Pull and enrich the latest Azure DevOps work items."""
    pat = _pat()
    if not pat:
        raise ValueError("AZDO_PAT not set — cannot pull live data")

    base = f"https://dev.azure.com/{ORG}/{PROJECT}/_apis"
    hdr = _auth_header(pat)
    query = {
        "query": (
            "SELECT [System.Id] FROM WorkItems "
            f"WHERE [System.TeamProject] = '{PROJECT}' "
            "AND [System.WorkItemType] <> '' "
            "ORDER BY [System.ChangedDate] DESC"
        )
    }
    resp = requests.post(f"{base}/wit/wiql?api-version={API_VERSION}",
                         headers=hdr, json=query, timeout=30)
    resp.raise_for_status()
    ids = [item["id"] for item in resp.json().get("workItems", [])]

    fields = [
        "System.Id", "System.Title", "System.WorkItemType", "System.State",
        "System.AssignedTo", "System.IterationPath", "System.AreaPath",
        "System.Parent", "Microsoft.VSTS.Scheduling.StoryPoints",
        "Microsoft.VSTS.Common.Priority", "System.CreatedDate",
        "System.ChangedDate", "System.Tags", "System.BoardColumn",
        "System.BoardColumnDone", "System.BoardLane",
    ]
    all_items = []
    for i in range(0, len(ids), 200):
        chunk = ids[i:i + 200]
        resp = requests.post(
            f"{base}/wit/workitemsbatch?api-version={API_VERSION}",
            headers=hdr, json={"ids": chunk, "fields": fields}, timeout=30)
        resp.raise_for_status()
        all_items.extend(resp.json().get("value", []))

    work_types = {
        item.get("fields", {}).get("System.WorkItemType")
        for item in all_items
        if item.get("fields", {}).get("System.WorkItemType")
    }
    categories = _state_categories(base, hdr, work_types)
    for item in all_items:
        item_fields = item.get("fields", {})
        item["_state_category"] = categories.get((
            item_fields.get("System.WorkItemType"),
            item_fields.get("System.State"),
        ))
    return all_items


def _category(state, category=None):
    if category:
        return category
    if state in DONE:
        return "Completed"
    if state in ACTIVE:
        return "InProgress"
    if str(state).lower() in {"removed", "deleted", "cut"}:
        return "Removed"
    return "Proposed"


def _wi(w):
    """Convert a raw Azure work item JSON to our normalized dict."""
    f = w.get("fields", {})
    state = f.get("System.State") or "Unknown"
    return {
        "id": w.get("id"),
        "title": (f.get("System.Title") or ""),
        "type": (f.get("System.WorkItemType") or "Unknown"),
        "state": state,
        "state_category": _category(state, w.get("_state_category")),
        "board_column": f.get("System.BoardColumn") or state,
        "board_column_done": bool(f.get("System.BoardColumnDone")),
        "board_lane": f.get("System.BoardLane") or "Default",
        "assignee": (f.get("System.AssignedTo") or {}).get("displayName")
        if isinstance(f.get("System.AssignedTo"), dict) else (f.get("System.AssignedTo") or "Unassigned"),
        "sprint": _sprint(f.get("System.IterationPath")),
        "area": _leaf(f.get("System.AreaPath") or PROJECT),
        "sp": f.get("Microsoft.VSTS.Scheduling.StoryPoints"),
        "priority": f.get("Microsoft.VSTS.Common.Priority"),
        "created": _parse_date(f.get("System.CreatedDate")),
        "changed": _parse_date(f.get("System.ChangedDate")),
        "parent": f.get("System.Parent"),
        "tags": [t.strip() for t in str(f.get("System.Tags") or "").split(";") if t.strip()],
        "url": f"https://dev.azure.com/{ORG}/{PROJECT}/_workitems/edit/{w.get('id')}",
    }


def read_workbook_items():
    """Fallback: read normalized items from the local workbook's Raw Data.
    Returns [] when the workbook is absent (e.g. on a cloud host)."""
    if not os.path.exists(WORKBOOK):
        return []
    wb = openpyxl.load_workbook(WORKBOOK, data_only=True)
    ws = wb["Raw Data"]
    hdrs = {}  # name -> index handled by positions below
    header = [c.value for c in ws[5]]
    idx = {name: i for i, name in enumerate(header) if name}
    items = []
    for row in ws.iter_rows(min_row=6, values_only=True):
        if not row or row[0] is None and (row[1] is None):
            continue
        def g(name, default=None):
            i = idx.get(name)
            return row[i] if i is not None and i < len(row) else default
        created = _parse_date(g("Created Date"))
        state = g("State") or "Unknown"
        items.append({
            "id": g("Work Item ID"),
            "title": g("Title") or "",
            "type": g("Work Item Type") or "Unknown",
            "state": state,
            "state_category": _category(state, g("State Category")),
            "board_column": g("Board Column") or state,
            "board_column_done": bool(g("Board Column Done")),
            "board_lane": g("Board Lane") or "Default",
            "assignee": g("Assigned To") or "Unassigned",
            "sprint": _sprint(g("Iteration Path")),
            "area": _leaf(g("Area Path") or PROJECT),
            "sp": g("Story Points"),
            "priority": g("Priority"),
            "created": created,
            "changed": _parse_date(g("Changed Date")),
            "parent": g("Parent ID"),
            "tags": [t.strip() for t in str(g("Tags") or "").split(";") if t.strip()],
            "url": f"https://dev.azure.com/{ORG}/{PROJECT}/_workitems/edit/{g('Work Item ID')}",
        })
    return items


def load_items(force_pull=True):
    """Try live Azure pull; fall back to workbook Raw Data.
    Returns (items, mode). mode in {'live','workbook','empty'}."""
    if force_pull and _pat():
        try:
            raw = pull_from_azure()
            items = [_wi(w) for w in raw]
            return items, "live"
        except Exception as exc:
            # keep the real error visible (for debugging), fall back to cache
            st.session_state["last_pull_error"] = str(exc)
    items = read_workbook_items()
    if items:
        return items, "workbook"
    return [], "empty"


# ---------------------------------------------------------------- analysis
def is_done(i):
    return i.get("state_category") == "Completed"


def is_open(i):
    return i.get("state_category") not in TERMINAL_CATEGORIES


def is_active(i):
    return i.get("state_category") == "InProgress"


def item_metrics(items):
    total = len(items)
    done = sum(is_done(i) for i in items)
    active = sum(is_active(i) for i in items)
    return {
        "total": total,
        "done": done,
        "active": active,
        "open": sum(is_open(i) for i in items),
        # User Stories are intentionally shared across task owners; only an
        # unassigned Task is an actionable ownership risk.
        "unassigned": sum(
            i["type"] == "Task" and i["assignee"] == "Unassigned"
            for i in items
        ),
        "stale": sum(is_open(i) and i["created"] and (dt.date.today() - i["created"]).days >= STALE_DAYS for i in items),
        "done_pct": done / total if total else 0,
    }


def type_progress(items):
    """Independent % per type with child->parent roll-up when Parent ID present."""
    children = defaultdict(list)
    for i in items:
        if i["parent"] is not None:
            children[i["parent"]].append(i)
    hier = any(i["parent"] is not None for i in items)

    memo = {}
    visiting = set()

    def effective(item):
        item_id = item["id"]
        if item_id in memo:
            return memo[item_id]
        if item_id in visiting:
            # Malformed Azure hierarchy cycle: do not count it as complete.
            return False
        visiting.add(item_id)
        child_items = children.get(item_id, [])
        result = (
            all(effective(child) for child in child_items)
            if child_items
            else is_done(item)
        )
        visiting.remove(item_id)
        memo[item_id] = result
        return result

    agg = defaultdict(lambda: {"total": 0, "done": 0})
    for t in DELIVERY_TYPES:
        agg[t] = {"total": 0, "done": 0}
    for i in items:
        d = effective(i) if hier else is_done(i)
        agg[i["type"]]["total"] += 1
        agg[i["type"]]["done"] += 1 if d else 0
    out = {}
    for t, m in agg.items():
        out[t] = {"total": m["total"], "done": m["done"],
                  "pct": m["done"] / m["total"] if m["total"] else None}
    out["hierarchy_used"] = hier
    return out


def scope_metrics(items):
    """Story Done when closed AND all child Tasks done (when hierarchy present)."""
    stories = [i for i in items if i["type"] == "User Story"]
    tasks = [i for i in items if i["type"] == "Task"]
    tasks_by_parent = defaultdict(list)
    hier = any(i["parent"] is not None for i in items)
    for t in tasks:
        if t["parent"] is not None:
            tasks_by_parent[t["parent"]].append(t)
    done_story_items = []
    for story in stories:
        child_tasks = tasks_by_parent.get(story["id"])
        # A parent with children rolls up from them. A story without linked
        # children falls back to its own Azure state.
        story_done = (
            all(is_done(task) for task in child_tasks)
            if child_tasks and hier
            else is_done(story)
        )
        if story_done:
            done_story_items.append(story)

    task_done_count = sum(is_done(task) for task in tasks)
    total_sp = sum(float(story["sp"] or 0) for story in stories)
    done_sp = sum(float(story["sp"] or 0) for story in done_story_items)
    return {
        "stories": len(stories),
        "stories_done": len(done_story_items),
        "scope_pct": len(done_story_items) / len(stories) if stories else None,
        "tasks": len(tasks),
        "tasks_done": task_done_count,
        "task_pct": task_done_count / len(tasks) if tasks else None,
        "total_sp": total_sp,
        "done_sp": done_sp,
        "velocity_pct": done_sp / total_sp if total_sp else None,
        "hier": hier,
    }


def percent(value):
    """Convert a 0..1 ratio to a display-ready numeric percentage."""
    return round(value * 100, 1) if value is not None else None


COLUMN_AR = {
    "Work Type": "نوع العنصر", "Total": "الإجمالي", "Done": "مكتمل",
    "Completion %": "نسبة الاكتمال", "State": "الحالة", "Category": "التصنيف",
    "Items": "العناصر", "Iteration": "الدورة", "Total Dev Items": "إجمالي عناصر التسليم",
    "User Stories": "قصص المستخدم", "Stories Done": "القصص المكتملة",
    "Scope Done %": "اكتمال النطاق", "Tasks": "المهام", "Tasks Done": "المهام المكتملة",
    "Task Done %": "اكتمال المهام", "Active": "قيد التنفيذ", "Unassigned": "بدون مسؤول",
    "ID": "المعرف", "Title": "العنوان", "Type": "النوع", "State Category": "تصنيف الحالة",
    "Board Column": "عمود اللوحة", "Board Lane": "مسار اللوحة", "Assignee": "المسؤول",
    "Area": "المجال", "Tags": "الوسوم", "Priority": "الأولوية", "Created": "تاريخ الإنشاء",
    "Changed": "تاريخ التعديل", "Age (d)": "العمر بالأيام", "Parent ID": "معرف الأصل",
    "Azure Link": "رابط Azure", "Tag": "الوسم", "Stories": "القصص", "Scope %": "النطاق",
    "Task %": "نسبة اكتمال المهام", "Areas": "المجالات", "Task Completion %": "اكتمال المهام",
    "Open": "مفتوح", "Stories Involved": "القصص المشاركة", "Stories Fully Done": "القصص المكتملة",
    "SP": "النقاط", "Done SP": "النقاط المكتملة", "Risk": "المخاطر", "Age": "العمر",
    "Sprint": "السبرينت", "Check": "الفحص", "Count": "العدد", "Interpretation": "التفسير",
    "Work Item ID": "معرف العنصر", "Work Item Type": "نوع العنصر", "Assigned To": "المسؤول",
    "Iteration Path": "مسار الدورة", "Area Path": "مسار المجال", "Story Points": "نقاط القصة",
    "Created Date": "تاريخ الإنشاء", "Changed Date": "تاريخ التعديل", "Board Column Done": "اكتمال عمود اللوحة",
    "Repository": "المستودع", "Repository ID": "معرف المستودع", "Default Branch": "الفرع الافتراضي",
    "Size (bytes)": "الحجم بالبايت", "Disabled": "معطّل", "Remote URL": "رابط الاستنساخ",
    "Contributor": "المساهم", "Email": "البريد", "Repositories": "المستودعات", "Commits": "Commits",
    "Pushes": "عمليات الرفع", "Pull Requests": "Pull Requests", "Commit ID": "معرف Commit",
    "Short ID": "المعرف المختصر", "Message": "الرسالة", "Author": "الكاتب", "Author Email": "بريد الكاتب",
    "Author Date": "تاريخ الكتابة", "Committer": "منفذ Commit", "Commit Date": "تاريخ Commit",
    "Change Counts": "ملخص التغييرات", "Push ID": "معرف الرفع", "Pushed By": "رفع بواسطة",
    "Pusher Email": "بريد من قام بالرفع", "Push Date": "تاريخ الرفع", "Branches": "الفروع",
    "PR ID": "معرف PR", "Description": "الوصف", "Status": "الحالة", "Draft": "مسودة",
    "Created By": "أنشأ بواسطة", "Creator Email": "بريد المنشئ", "Closed Date": "تاريخ الإغلاق",
    "Source Branch": "فرع المصدر", "Target Branch": "الفرع المستهدف", "Reviewers": "المراجعون",
    "Merge Status": "حالة الدمج", "Merge Commit": "Commit الدمج", "Change #": "رقم التغيير",
    "Change Type": "نوع التغيير", "Path": "المسار", "Original Path": "المسار الأصلي",
    "Git Object Type": "نوع Git", "Object ID": "معرف الكائن", "Operation": "العملية", "Error": "الخطأ",
}


def column_label(name):
    return COLUMN_AR.get(name, name) if is_ar else name


def localized_frame(frame):
    if not is_ar:
        return frame
    return frame.rename(columns={name: column_label(name) for name in frame.columns})


def percentage_columns(*names):
    """Streamlit table formatting for numeric 0..100 percentage columns."""
    return {
        column_label(name): st.column_config.NumberColumn(column_label(name), format="%.1f%%")
        for name in names
    }


def ribbon(scope, task, unassigned, stale):
    red = 0
    if (scope or 0) < 0.4:
        red += 1
    if (task or 0) < 0.4:
        red += 1
    if unassigned and unassigned >= 25:
        red += 1
    if stale and stale >= 10:
        red += 1
    if red >= 3:
        return "CRITICAL", dashboard_theme.ACCENTS["red"]
    if red >= 2:
        return "AT RISK", dashboard_theme.ACCENTS["amber"]
    return "HEALTHY", dashboard_theme.ACCENTS["green"]


def delivery_action(items, unassigned):
    open_count = sum(is_open(i) for i in items)
    if unassigned >= 25 and open_count:
        return tr(
            f"{unassigned} tasks are unassigned — assign owners first",
            f"يوجد {unassigned} مهمة بدون مسؤول — ابدأ بتحديد المسؤولين",
        )
    scope = scope_metrics(items)
    if items and scope["scope_pct"] == 0 and open_count:
        return tr(
            "Scope stalled — no fully-complete story yet",
            "النطاق متعطل — لا توجد قصة مكتملة بالكامل حتى الآن",
        )
    if any(i["type"] == "User Story" and i["sp"] is None for i in items):
        return tr(
            "Add Story Points to unestimated stories",
            "أضف Story Points للقصص غير المقدّرة",
        )
    return tr("Delivery on track", "التسليم يسير حسب الخطة")


# ============================================================ UI HELPERS
def tr(english, arabic):
    return arabic if is_ar else english


def theme_chart(chart):
    """Apply the active Altair theme to a chart before rendering."""
    return chart


def kpi_card(column, label, value, accent, icon="◆", subcaption=None):
    with column:
        sub_html = f"<div class='kpi-sub'>{subcaption}</div>" if subcaption else ""
        st.markdown(
            f"<div class='kpi-card' style='--accent:{accent}'>"
            f"<div class='kpi-card-top'><span class='kpi-label'>{label}</span>"
            f"<span class='kpi-icon'>{icon}</span></div>"
            f"<div class='kpi-value'>{value}</div>{sub_html}</div>",
            unsafe_allow_html=True,
        )


def localized_label(name):
    return column_label(name)


def weekly_creation_closure(delivery_items):
    """Weekly created vs closed item counts from Azure dates."""
    created = Counter()
    closed = Counter()
    for item in delivery_items:
        if not item.get("created"):
            continue
        monday = item["created"] - dt.timedelta(days=item["created"].weekday())
        created[monday] += 1
        closure = item.get("changed") if is_done(item) else None
        if closure:
            closed_monday = closure - dt.timedelta(days=closure.weekday())
            closed[closed_monday] += 1
    weeks = sorted(set(created) | set(closed))
    return pd.DataFrame([
        {
            "period": week.strftime("%d %b"),
            "Created": created.get(week, 0),
            "Closed": closed.get(week, 0),
        }
        for week in weeks[-12:]
    ])




# ---------------------------------------------------------------- rendering
st.set_page_config(
    page_title="Delivery Manager — Hoteliana",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
)

is_ar = st.session_state.get("language_selector", "العربية") == "العربية"
theme_mode = "dark" if st.session_state.get("theme_selector", "Dark") == "Dark" else "light"
apply_theme(theme_mode, is_ar)
chart_theme = dashboard_theme.chart_theme(theme_mode)
ACCENT = dashboard_theme.ACCENTS

st.sidebar.markdown(
    """
    <div class="sidebar-brand">
      <div class="brand-mark">M</div>
      <div><strong>MATN DELIVERY</strong><span>Hoteliana workspace</span></div>
    </div>
    """,
    unsafe_allow_html=True,
)
st.sidebar.segmented_control(
    tr("Language", "اللغة"),
    ["العربية", "English"],
    key="language_selector",
    default="العربية",
    label_visibility="collapsed",
)
st.sidebar.segmented_control(
    tr("Theme", "المظهر"),
    ["Dark", "Light"],
    key="theme_selector",
    default="Dark",
    label_visibility="collapsed",
)
st.sidebar.markdown(
    f'<div class="sidebar-label">{tr("DATA CONTROL", "التحكم بالبيانات")}</div>',
    unsafe_allow_html=True,
)

st.markdown(
    f"""
    <section class="enterprise-header">
      <div>
        <div class="header-kicker">MATN SOLUTIONS</div>
        <h1>{tr("Strategic Delivery Command Center", "مركز القيادة الاستراتيجي للتسليم")}</h1>
      </div>
      <div class="azure-pill"><span class="azure-dot"></span>Azure DevOps · {PROJECT}</div>
    </section>
    """,
    unsafe_allow_html=True,
)

user_missing = not _pat()
if user_missing:
    st.sidebar.warning(tr(
        "AZDO_PAT is not set — showing the latest workbook snapshot.",
        "رمز AZDO_PAT غير مضاف — يتم عرض آخر نسخة محفوظة من البيانات.",
    ))

refresh = st.sidebar.button(
    tr("↻  Sync Azure DevOps", "↻  مزامنة Azure DevOps"),
    type="primary",
)

if refresh:
    st.session_state.pop("azure_items", None)
    st.session_state.pop("azure_data_mode", None)

if "azure_items" not in st.session_state:
    loading_text = tr(
        "Syncing Azure DevOps data...",
        "جارٍ مزامنة بيانات Azure DevOps...",
    ) if _pat() else tr(
        "Loading saved data...",
        "جارٍ تحميل البيانات المحفوظة...",
    )
    with st.spinner(loading_text):
        loaded_items, loaded_mode = load_items(force_pull=bool(_pat()))
    st.session_state["azure_items"] = loaded_items
    st.session_state["azure_data_mode"] = loaded_mode

items = st.session_state["azure_items"]
data_mode = st.session_state["azure_data_mode"]
if refresh:
    if data_mode == "live":
        st.sidebar.success(tr(
            f"Synced {len(items)} items from Azure.",
            f"تمت مزامنة {len(items)} عنصر من Azure.",
        ))
    else:
        error = st.session_state.get("last_pull_error", "Azure credentials are unavailable.")
        st.sidebar.error(tr(
            f"Sync failed; saved data is shown. {error}",
            f"فشلت المزامنة؛ يتم عرض البيانات المحفوظة. {error}",
        ))

dev = [i for i in items if i["type"] in DEV_TYPES]

if not dev:
    st.sidebar.warning("No data loaded.")
    st.error("No dev work items found.")
    missing_pat = not _pat()
    if missing_pat:
        st.info("`AZDO_PAT` secret is not set. Set it in the hosting platform's secrets "
                "(Streamlit Cloud → Settings → Secrets) or as the GitHub Actions secret, "
                "then press **Refresh from Azure DevOps**.")
    # Show the underlying pull failure so it isn't a silent "nothing on screen"
    if st.session_state.get("last_pull_error"):
        st.warning(f"Last Azure pull failed: {st.session_state['last_pull_error']}")
    st.info("If you are self-hosting without a workbook, the app needs live Azure access "
            "via AZDO_PAT to load any data.")
    st.stop()

# Precompute shared analysis once (executive-only metrics used by sidebar).
all_m = item_metrics(dev)
scope = scope_metrics([i for i in dev if i["sprint"] != PB])
verdict, color = ribbon(scope["scope_pct"], scope["task_pct"], all_m["unassigned"], all_m["stale"])
prog = type_progress(dev)

# ---- Sidebar navigation: one section per tab, mirroring the Excel workbook
PAGES = {
    "Executive Dashboard": ("◈  Executive overview", "◈  النظرة التنفيذية"),
    "Sprint Summary": ("⏱  Sprint summary", "⏱  ملخص السبرينت"),
    "Sprint Board": ("▦  Sprint board", "▦  لوحة السبرينت"),
    "Tag Analysis": ("#  Tag analysis", "#  تحليل الوسوم"),
    "Team Analysis": ("👥  Team delivery", "👥  أداء الفريق"),
    "Area Analysis": ("◇  Area analysis", "◇  تحليل المجالات"),
    "Active Now": ("⚡  Active now", "⚡  العمل الحالي"),
    "Risks & Aging": ("⚠  Risks & aging", "⚠  المخاطر والتقادم"),
    "Data Quality": ("✓  Data quality", "✓  جودة البيانات"),
    "Repository Intelligence": ("📊  Repository intelligence", "📊  ذكاء المستودعات"),
    "Releases": ("🏆  Releases", "🏆  الإصدارات"),
    "Raw Data": ("📁  Raw data", "📁  البيانات الخام"),
}
# Grouped purely for sidebar presentation — every PAGES key appears in exactly
# one group. Streamlit has no native nested/grouped radio, so each group is
# its own st.sidebar.radio; selecting any of them writes the shared
# "dashboard_page" key that the dispatch block below reads.
PAGE_GROUPS = [
    (("Overview", "نظرة عامة"), ["Executive Dashboard", "Sprint Summary", "Sprint Board"]),
    (("Delivery & Team", "التسليم والفريق"), ["Tag Analysis", "Team Analysis", "Area Analysis", "Active Now"]),
    (("Quality & Risk", "الجودة والمخاطر"), ["Risks & Aging", "Data Quality"]),
    (("Data & Repositories", "البيانات والمستودعات"), ["Repository Intelligence", "Releases", "Raw Data"]),
]
assert sorted(sum((keys for _, keys in PAGE_GROUPS), [])) == sorted(PAGES), \
    "PAGE_GROUPS must partition PAGES exactly"

if "dashboard_page" not in st.session_state:
    st.session_state["dashboard_page"] = next(iter(PAGES))


def _sync_dashboard_page(group_key):
    st.session_state["dashboard_page"] = st.session_state[group_key]


for (label_en, label_ar), group_keys in PAGE_GROUPS:
    st.sidebar.markdown(
        f'<div class="sidebar-label">{tr(label_en, label_ar)}</div>',
        unsafe_allow_html=True,
    )
    group_key = f"nav_group_{label_en}"
    current = st.session_state["dashboard_page"]
    if group_key not in st.session_state:
        st.session_state[group_key] = current if current in group_keys else group_keys[0]
    st.sidebar.radio(
        label_en,
        group_keys,
        format_func=lambda key: PAGES[key][1 if is_ar else 0],
        key=group_key,
        label_visibility="collapsed",
        on_change=_sync_dashboard_page,
        args=(group_key,),
    )

page = st.session_state["dashboard_page"]
st.sidebar.markdown(
    f'<div class="sidebar-label">{tr("SCOPE SNAPSHOT", "ملخص النطاق")}</div>',
    unsafe_allow_html=True,
)
source_col, scope_col = st.sidebar.columns(2)
source_col.metric(tr("Source", "المصدر"), tr(data_mode.title(), "مباشر" if data_mode == "live" else "محفوظ"))
scope_col.metric(tr("Delivery", "التسليم"), len(dev))
st.sidebar.caption(tr(
    f"{len(items):,} total Azure work items",
    f"إجمالي عناصر Azure: {len(items):,}",
))

st.caption(tr(
    f"Live synchronization with {ORG}/{PROJECT}." if data_mode == "live" else "Showing the latest saved Azure snapshot.",
    f"مزامنة مباشرة مع {ORG}/{PROJECT}." if data_mode == "live" else "يتم عرض آخر نسخة محفوظة من بيانات Azure.",
))
st.divider()


# ============================================================ EXECUTIVE
def render_executive():
    st.header(tr("Executive Delivery Overview", "النظرة التنفيذية للتسليم"))
    st.caption(tr(
        "Live Azure DevOps delivery health, ownership and execution signals.",
        "مؤشرات حية لصحة التسليم والتنفيذ وتوزيع المسؤوليات من Azure DevOps.",
    ))
    verdict_label = {
        "HEALTHY": tr("Healthy", "مستقر"),
        "AT RISK": tr("At risk", "معرّض للخطر"),
        "CRITICAL": tr("Critical", "حرج"),
    }[verdict]
    st.markdown(
        f"<div class='health-ribbon' style='--health:{color}'>"
        f"{verdict_label} · {delivery_action(dev, all_m['unassigned'])}</div>",
        unsafe_allow_html=True,
    )

    if not dev:
        st.warning(tr("No work items to chart yet.", "لا توجد عناصر عمل لعرضها."))
        return

    columns = st.columns(6, gap="small")
    kpi_card(columns[0], tr("Story scope done", "نطاق القصص المكتمل"), f"{scope['scope_pct'] or 0:.0%}", ACCENT["green"],
              icon="%", subcaption=tr(f"{scope['stories_done']} of {scope['stories']} stories", f"{scope['stories_done']} من {scope['stories']} قصة"))
    kpi_card(columns[1], tr("Task completion", "اكتمال المهام"), f"{scope['task_pct'] or 0:.0%}", ACCENT["blue"],
              icon="✓", subcaption=tr(f"{scope['tasks_done']} of {scope['tasks']} tasks", f"{scope['tasks_done']} من {scope['tasks']} مهمة"))
    kpi_card(columns[2], tr("Active now", "قيد التنفيذ"), all_m["active"], ACCENT["purple"],
              icon="⚡", subcaption=tr("items in progress", "عنصر نشط الآن"))
    kpi_card(columns[3], tr("Unassigned", "بدون مسؤول"), all_m["unassigned"], ACCENT["red"],
              icon="!", subcaption=tr("tasks need an owner", "مهمة تحتاج تعيين"))
    backlog_count = sum(1 for item in dev if item["sprint"] == PB)
    kpi_card(columns[4], tr("Product backlog", "قائمة المنتج"), backlog_count, ACCENT["gold"],
              icon="▤", subcaption=tr("items with no sprint", "عنصر بدون سبرينت"))
    kpi_card(columns[5], tr("Stale ≥14d", "متقادم ≥14 يوم"), all_m["stale"], ACCENT["amber"],
              icon="⏱", subcaption=tr("no delay" if not all_m["stale"] else "needs attention", "لا يوجد تأخير" if not all_m["stale"] else "يحتاج متابعة"))

    section_header(
        "Completion by work type",
        "الاكتمال حسب نوع عنصر العمل", "◈")
    prog_df = pd.DataFrame([
        {"Work Type": work_type, "Total": prog[work_type]["total"],
         "Done": prog[work_type]["done"],
         "Completion %": percent(prog[work_type]["pct"])}
        for work_type in DELIVERY_TYPES
    ])
    table_col, chart_col = st.columns([1, 1.15])
    with table_col:
        st.dataframe(
            localized_frame(prog_df), width="stretch", hide_index=True,
            column_config={
                **percentage_columns("Completion %"),
                localized_label("Total"): st.column_config.NumberColumn(
                    localized_label("Total"), format="%d"),
                localized_label("Done"): st.column_config.NumberColumn(
                    localized_label("Done"), format="%d"),
            },
        )
    with chart_col:
        type_colors = {
            "Epic": ACCENT["purple"], "Feature": ACCENT["blue"],
            "User Story": ACCENT["teal"], "Task": ACCENT["green"], "Bug": ACCENT["red"],
        }
        st.altair_chart(dashboard_theme.donut_chart(
            prog_df.rename(columns={"Work Type": "kind", "Total": "total"}),
            "kind", "total", chart_theme, colors=type_colors,
        ), width="stretch")
    hier_txt = tr(
        "child→parent roll-up active ✅",
        "تجميع نتائج الأبناء إلى العناصر الرئيسية مفعّل ✅",
    ) if prog["hierarchy_used"] else tr(
        "own-state (add Parent ID for roll-up)",
        "الاعتماد على حالة العنصر (أضف Parent ID لتفعيل التجميع)",
    )
    st.caption(tr(f"Hierarchy: {hier_txt}", f"التسلسل الهرمي: {hier_txt}"))

    section_header("Current board flow", "تدفق العمل الحالي", "▦")
    state_df = pd.DataFrame([
        {"State": state, "Category": category, "Items": count}
        for (state, category), count in Counter(
            (i["state"], i["state_category"]) for i in dev
        ).most_common()
    ])
    category_colors = {
        "Completed": ACCENT["green"], "InProgress": ACCENT["blue"],
        "Resolved": ACCENT["teal"], "Proposed": ACCENT["gold"], "Removed": ACCENT["red"],
    }
    table_col, chart_col = st.columns([1, 1.2])
    with table_col:
        st.caption(tr("Exact Azure states and canonical categories", "حالات Azure الفعلية وتصنيفاتها"))
        st.dataframe(localized_frame(state_df), width="stretch", hide_index=True)
    with chart_col:
        st.caption(tr("Items by Azure Board column and category", "العناصر حسب عمود اللوحة والتصنيف"))
        board_flow = pd.DataFrame([
            {"column": column_name, "category": category, "items": count}
            for (column_name, category), count in Counter(
                (i["board_column"], i["state_category"]) for i in dev
            ).most_common()
        ])
        st.altair_chart(dashboard_theme.stacked_hbar_chart(
            board_flow, "column", "category", "items", chart_theme
        ), width="stretch")

    section_header("Delivery momentum", "زخم التسليم", "⚡")
    week_frame = weekly_creation_closure(dev)
    if week_frame.empty:
        st.info(tr("No dated items yet.", "لا توجد عناصر بتواريخ بعد."))
    else:
        st.altair_chart(dashboard_theme.area_trend_chart(
            week_frame, "period", ("Closed", "Created"),
            {"Closed": ACCENT["green"], "Created": ACCENT["blue"]}, chart_theme,
        ), width="stretch")


# ============================================================ SPRINT SUMMARY
def sprint_summary_df():
    groups = defaultdict(list)
    for i in dev:
        groups[i["sprint"]].append(i)
    rows = []
    for sprint in sorted(groups, key=lambda s: (s == PB, s)):
        m = groups[sprint]
        scg = scope_metrics(m)
        im = item_metrics(m)
        rows.append({
            "Iteration": sprint,
            "Total Dev Items": len(m),
            "User Stories": scg["stories"],
            "Stories Done": scg["stories_done"],
            "Scope Done %": percent(scg["scope_pct"]),
            "Tasks": scg["tasks"],
            "Tasks Done": scg["tasks_done"],
            "Task Done %": percent(scg["task_pct"]),
            "Active": im["active"],
            "Unassigned": im["unassigned"],
            f"Open ≥{STALE_DAYS}d": im["stale"],
            "Iteration Meaning": "No sprint assigned" if sprint == PB else "Committed iteration",
        })
    return pd.DataFrame(rows)


def render_sprint_summary():
    st.header(tr("Sprint Summary", "ملخص السبرينت"))
    st.caption(tr("One row per Azure iteration; Product Backlog is shown separately.", "صف لكل دورة Azure مع عرض Product Backlog بشكل منفصل."))
    summary = sprint_summary_df()
    chart_col, table_col = st.columns([1, 1.25])
    with chart_col:
        section_header("Stories done vs total", "القصص المكتملة مقابل الإجمالي", "◷")
        story_chart = summary.rename(columns={
            "Iteration": "iteration", "Stories Done": "done", "User Stories": "total",
        })
        st.altair_chart(dashboard_theme.grouped_hbar_chart(
            story_chart, "iteration", ("done", "total"),
            {"done": ACCENT["green"], "total": ACCENT["blue"]}, chart_theme,
        ), width="stretch")
    with table_col:
        st.dataframe(
            localized_frame(summary), width="stretch", hide_index=True,
            column_config=percentage_columns("Scope Done %", "Task Done %"),
        )


# ============================================================ SPRINT BOARD
def render_sprint_board():
    st.header(tr("Sprint Board", "لوحة السبرينت"))
    st.caption(tr("Delivery work grouped by iteration and assignee, with Azure links.", "عناصر التسليم مجمعة حسب الدورة والمسؤول مع روابط Azure."))
    df = pd.DataFrame([{
        "Iteration": i["sprint"], "ID": i["id"], "Title": i["title"], "Type": i["type"],
        "State": i["state"], "State Category": i["state_category"],
        "Board Column": i["board_column"], "Board Lane": i["board_lane"],
        "Assignee": i["assignee"], "Area": i["area"],
        "Tags": "; ".join(i["tags"]) or "Untagged", "SP": i["sp"],
        "Priority": i["priority"],
        "Created": i["created"], "Changed": i["changed"],
        "Age (d)": (dt.date.today() - i["created"]).days if i["created"] else None,
        "Parent ID": i["parent"], "Azure Link": i["url"],
    } for i in dev])
    st.dataframe(localized_frame(df), width="stretch", hide_index=True, height=500)


# ============================================================ TAG ANALYSIS
def render_tag_analysis():
    st.header(tr("Tag Analysis", "تحليل الوسوم"))
    st.caption(tr("Multi-tag items are counted once per tag; untagged work is explicit.", "يتم احتساب العنصر تحت كل وسم مع إظهار العناصر غير المصنفة."))
    tag_map = defaultdict(list)
    for i in dev:
        for t in (i["tags"] or ["Untagged"]):
            tag_map[t].append(i)
    rows = []
    for tag, members in sorted(tag_map.items(), key=lambda x: -len(x[1])):
        sc = scope_metrics(members)
        im = item_metrics(members)
        rows.append({
            "Tag": tag, "Items": len(members),
            "Stories": sc["stories"], "Stories Done": sc["stories_done"],
            "Scope %": percent(sc["scope_pct"]),
            "Tasks": sc["tasks"], "Tasks Done": sc["tasks_done"],
            "Task %": percent(sc["task_pct"]),
            "Active": im["active"], "Unassigned": im["unassigned"],
            f"Open ≥{STALE_DAYS}d": im["stale"],
            "Areas": ", ".join(sorted({i["area"] for i in members})),
        })
    tag_frame = pd.DataFrame(rows)
    chart_col, table_col = st.columns([1, 1.3])
    with chart_col:
        section_header("Items per tag", "العناصر لكل وسم", "#")
        if not tag_frame.empty:
            tag_counts = tag_frame.rename(columns={"Tag": "tag", "Items": "items"})[
                ["tag", "items"]
            ]
            st.altair_chart(dashboard_theme.hbar_chart(
                tag_counts, "items", "tag", chart_theme, color=ACCENT["teal"],
            ), width="stretch")
    with table_col:
        st.dataframe(
            localized_frame(tag_frame), width="stretch", hide_index=True,
            column_config=percentage_columns("Scope %", "Task %"),
        )


# ============================================================ TEAM ANALYSIS
def team_df():
    """Task-only ownership metrics; stories are shared through child Tasks."""
    tasks = [item for item in dev if item["type"] == "Task"]
    tasks_by_assignee = defaultdict(list)
    tasks_by_story = defaultdict(list)
    for task in tasks:
        tasks_by_assignee[task["assignee"]].append(task)
        if task["parent"] is not None:
            tasks_by_story[task["parent"]].append(task)

    completed_story_ids = {
        story_id
        for story_id, child_tasks in tasks_by_story.items()
        if child_tasks and all(is_done(task) for task in child_tasks)
    }

    rows = []
    ordered_groups = sorted(
        tasks_by_assignee.items(), key=lambda group: (-len(group[1]), group[0])
    )
    for assignee, owned_tasks in ordered_groups:
        done_count = sum(is_done(task) for task in owned_tasks)
        involved_story_ids = {
            task["parent"] for task in owned_tasks if task["parent"] is not None
        }
        rows.append({
            "Assignee": assignee,
            "Tasks": len(owned_tasks),
            "Tasks Done": done_count,
            "Task Completion %": percent(done_count / len(owned_tasks)),
            "Active": sum(is_active(task) for task in owned_tasks),
            "Open": sum(is_open(task) for task in owned_tasks),
            f"Open ≥{STALE_DAYS}d": item_metrics(owned_tasks)["stale"],
            "Stories Involved": len(involved_story_ids),
            "Stories Fully Done": len(involved_story_ids & completed_story_ids),
            "Areas": ", ".join(sorted({task["area"] for task in owned_tasks})),
        })
    return pd.DataFrame(rows)


def render_team_analysis():
    st.header(tr("Team Delivery", "أداء الفريق"))
    st.caption(tr("Team contribution is measured through completed Tasks.", "تُقاس مساهمة أعضاء الفريق من خلال المهام المكتملة."))
    team = team_df()
    if not team.empty:
        table_col, chart_col = st.columns([1, 1.4])
        with table_col:
            st.dataframe(
                localized_frame(team), width="stretch", hide_index=True,
                column_config=percentage_columns("Task Completion %"),
            )
        with chart_col:
            section_header("Tasks per member", "المهام لكل عضو", "◎")
            member_load = team.rename(columns={"Assignee": "member", "Tasks": "tasks"})[
                ["member", "tasks"]
            ]
            st.altair_chart(dashboard_theme.hbar_chart(
                member_load, "tasks", "member", chart_theme,
                color=ACCENT["purple"],
            ), width="stretch")
    else:
        st.dataframe(localized_frame(team), width="stretch", hide_index=True)


# ============================================================ AREA ANALYSIS
def area_df():
    rows = []
    for area in sorted({i["area"] for i in dev}, key=lambda a: -sum(1 for i in dev if i["area"] == a)):
        members = [i for i in dev if i["area"] == area]
        sc = scope_metrics(members)
        im = item_metrics(members)
        rows.append({
            "Area": area, "Total": len(members),
            "Stories": sc["stories"], "Stories Done": sc["stories_done"],
            "Scope %": percent(sc["scope_pct"]),
            "Tasks": sc["tasks"], "Tasks Done": sc["tasks_done"],
            "Task %": percent(sc["task_pct"]),
            "SP": sc["total_sp"], "Done SP": sc["done_sp"],
            "Active": im["active"], "Unassigned": im["unassigned"],
            f"Open ≥{STALE_DAYS}d": im["stale"],
        })
    return pd.DataFrame(rows)


def render_area_analysis():
    st.header(tr("Area Analysis", "تحليل المجالات"))
    st.caption(tr("Scope and execution across Azure Area Paths.", "توزيع النطاق والتنفيذ حسب مسارات Azure."))
    areas = area_df()
    if not areas.empty:
        table_col, chart_col = st.columns([1, 1.4])
        with table_col:
            st.dataframe(
                localized_frame(areas), width="stretch", hide_index=True,
                column_config=percentage_columns("Scope %", "Task %"),
            )
        with chart_col:
            section_header("Delivery items per area", "عناصر التسليم لكل مجال", "◇")
            area_load = areas.rename(columns={"Area": "area", "Total": "total"})[
                ["area", "total"]
            ]
            st.altair_chart(dashboard_theme.hbar_chart(
                area_load, "total", "area", chart_theme, color=ACCENT["gold"],
            ), width="stretch")
    else:
        st.dataframe(localized_frame(areas), width="stretch", hide_index=True)


# ============================================================ ACTIVE NOW
def render_active_now():
    st.header(tr("Active & Open Work", "العمل الحالي والمفتوح"))
    st.caption(tr("Open work ordered by ownership and aging risk.", "العمل المفتوح مرتب حسب المسؤولية وخطر التقادم."))
    open_items = [i for i in dev if is_open(i)]
    def priority(item):
        if item["assignee"] == "Unassigned":
            return "Critical"
        if is_active(item):
            return "Doing"
        if item["created"] and (dt.date.today() - item["created"]).days >= STALE_DAYS:
            return "Aging"
        if item["type"] == "User Story":
            return "Scope"
        return "High"
    rows = []
    for i in sorted(open_items, key=lambda x: priority(x)):
        rows.append({
            "Priority": priority(i), "ID": i["id"], "Title": i["title"], "Type": i["type"],
            "State": i["state"], "Board Column": i["board_column"],
            "Assignee": i["assignee"], "Sprint": i["sprint"],
            "Area": i["area"], "Tags": "; ".join(i["tags"]) or "Untagged",
            "Age (d)": (dt.date.today() - i["created"]).days if i["created"] else None,
            "Azure Link": i["url"],
        })
    if rows:
        st.dataframe(localized_frame(pd.DataFrame(rows)), width="stretch", hide_index=True, height=500)
    else:
        st.success(tr("No open work — all done!", "لا يوجد عمل مفتوح — كل العناصر مكتملة!"))


# ============================================================ RISKS & AGING
def render_risks():
    st.header(tr("Risks & Aging", "المخاطر والتقادم"))
    st.caption(tr("Open work ranked by days since creation.", "العمل المفتوح مرتب حسب عدد الأيام منذ الإنشاء."))
    open_items = [i for i in dev if is_open(i)]
    ordered = sorted(open_items, key=lambda i: (i["assignee"] != "Unassigned",
                                                -(i["created"] and (dt.date.today() - i["created"]).days or -1),
                                                i["id"]))
    rows = []
    for i in ordered:
        risk = []
        if i["assignee"] == "Unassigned":
            risk.append("Unassigned")
        if i["sprint"] == PB:
            risk.append("No sprint")
        if i["created"] and (dt.date.today() - i["created"]).days >= STALE_DAYS:
            risk.append(f"Age ≥{STALE_DAYS}d")
        if i["type"] == "User Story" and i["sp"] is None:
            risk.append("No SP")
        age = (dt.date.today() - i["created"]).days if i["created"] else None
        rows.append({
            "Risk": ", ".join(risk) or "Monitor", "Age": age, "ID": i["id"], "Title": i["title"],
            "Type": i["type"], "State": i["state"], "Assignee": i["assignee"],
            "Sprint": i["sprint"], "Area": i["area"], "Tags": "; ".join(i["tags"]) or "Untagged",
            "Azure Link": i["url"],
        })
    st.dataframe(localized_frame(pd.DataFrame(rows)), width="stretch", hide_index=True, height=500)


# ============================================================ DATA QUALITY
def render_data_quality():
    st.header(tr("Data Quality & Trust", "جودة وموثوقية البيانات"))
    rows = [
        ("All Azure work items imported", len(items), "Exact"),
        ("Delivery scope (Epic+Feature+Story+Task+Bug)", len(dev), "Exact"),
        ("Test artifacts retained in Raw Data", len(items) - len(dev), "Excluded from delivery scope"),
        ("Items in real sprint paths", sum(1 for i in dev if i["sprint"] != PB), "Exact"),
        ("Product Backlog (no sprint)", sum(1 for i in dev if i["sprint"] == PB), "Exact"),
        ("Items without assignee", all_m["unassigned"], "Exact"),
        ("Items without tags", sum(1 for i in dev if not i["tags"]), "Exact"),
        ("User Stories without Story Points", sum(1 for i in dev if i["type"] == "User Story" and i["sp"] is None), "Exact"),
        ("Items with Parent ID", sum(1 for i in dev if i["parent"] is not None), "Enables hierarchy roll-up"),
    ]
    st.dataframe(localized_frame(pd.DataFrame(rows, columns=["Check", "Count", "Interpretation"])),
                 width="stretch", hide_index=True)


# ============================================================ REPOSITORY INTELLIGENCE
def _repository_rows(rows, selected_repositories):
    return [row for row in rows if row.get("Repository") in selected_repositories]


def _contributor_filter(rows, contributor, name_key):
    if contributor == "All":
        return rows
    return [row for row in rows if row.get(name_key) == contributor]


def _activity_table(title_en, title_ar, rows, height=420):
    st.subheader(tr(title_en, title_ar))
    if not rows:
        st.info(tr("No matching records.", "لا توجد سجلات مطابقة."))
        return
    st.dataframe(
        localized_frame(pd.DataFrame(rows)),
        width="stretch",
        hide_index=True,
        height=height,
    )


def _load_repository_activity():
    pat = _pat()
    if not pat:
        return None
    client = AzureRepoActivityClient(ORG, PROJECT, pat)
    return client.fetch_all()


def _clear_repository_cache_when_requested():
    label = tr("↻ Reload full repository history", "↻ إعادة تحميل كل تاريخ المستودعات")
    if not st.button(label):
        return
    st.session_state.pop("repository_activity", None)
    st.session_state.pop("repository_activity_error", None)


def _fetch_repository_cache():
    if "repository_activity" in st.session_state or not _pat():
        return
    message = tr(
        "Loading all repository history and file changes. This can take several minutes...",
        "جارٍ تحميل كل تاريخ المستودعات وتغييرات الملفات. قد يستغرق ذلك عدة دقائق...",
    )
    with st.spinner(message):
        try:
            st.session_state["repository_activity"] = _load_repository_activity()
        except requests.RequestException as exc:
            st.session_state["repository_activity_error"] = str(exc)


def _repository_cache():
    if not _pat():
        st.warning(tr(
            "AZDO_PAT with Code (Read) permission is required to load repository history.",
            "يلزم AZDO_PAT بصلاحية Code (Read) لتحميل تاريخ المستودعات.",
        ))
        return None
    error = st.session_state.get("repository_activity_error")
    if error:
        st.error(tr(
            f"Repository history could not be loaded: {error}",
            f"تعذر تحميل تاريخ المستودعات: {error}",
        ))
        return None
    activity = st.session_state.get("repository_activity")
    if not activity:
        st.info(tr("No repository activity loaded.", "لم يتم تحميل نشاط المستودعات."))
    return activity


def _repository_filter_controls(activity):
    repository_names = sorted(row["Repository"] for row in activity["repositories"])
    selected_repositories = st.multiselect(
        tr("Repositories", "المستودعات"), repository_names, default=repository_names
    )
    contributor_names = {
        row.get(key)
        for dataset_key, key in (("commits", "Author"), ("pushes", "Pushed By"),
                                 ("pull_requests", "Created By"))
        for row in activity[dataset_key] if row.get(key)
    }
    contributor = st.selectbox(
        tr("Contributor", "المساهم"), ["All", *sorted(contributor_names)],
        format_func=lambda value: tr(value, "الكل") if value == "All" else value,
    )
    statuses = sorted({row.get("Status", "") for row in activity["pull_requests"]})
    selected_statuses = st.multiselect(
        tr("Pull request status", "حالة Pull Request"), statuses, default=statuses
    )
    return selected_repositories, contributor, selected_statuses


def _filtered_repository_activity(activity, filter_values):
    repositories, contributor, statuses = filter_values
    selected = {
        key: _repository_rows(activity[key], repositories)
        for key in ("repositories", "commits", "pushes", "pull_requests", "changes")
    }
    selected["commits"] = _contributor_filter(selected["commits"], contributor, "Author")
    selected["pushes"] = _contributor_filter(selected["pushes"], contributor, "Pushed By")
    selected["pull_requests"] = _contributor_filter(
        selected["pull_requests"], contributor, "Created By"
    )
    selected["pull_requests"] = [
        row for row in selected["pull_requests"] if row.get("Status") in statuses
    ]
    selected["changes"] = _contributor_filter(selected["changes"], contributor, "Author")
    selected["contributors"] = contributor_rows(selected)
    return selected


def _render_repository_kpis(activity):
    columns = st.columns(6, gap="small")
    values = (
        (tr("Repositories", "المستودعات"), len(activity["repositories"]), ACCENT["blue"], "⌘"),
        (tr("Contributors", "المساهمون"), len(activity["contributors"]), ACCENT["purple"], "◎"),
        ("Commits", len(activity["commits"]), ACCENT["green"], "✓"),
        (tr("Pushes", "عمليات الرفع"), len(activity["pushes"]), ACCENT["teal"], "↻"),
        ("Pull Requests", len(activity["pull_requests"]), ACCENT["gold"], "⌥"),
        (tr("Changed files", "الملفات المتغيرة"), len(activity["changes"]), ACCENT["pink"], "▤"),
    )
    for column, (label, value, accent, icon) in zip(columns, values):
        kpi_card(column, label, value, accent, icon=icon)


def _render_repository_tables(activity, failures):
    table_col, chart_col = st.columns([1, 1.4])
    with table_col:
        _activity_table("Contributor summary", "ملخص المساهمين", activity["contributors"])
    with chart_col:
        section_header("Commits per contributor", "Commits لكل مساهم", "⌘")
        if activity["contributors"]:
            commit_load = [
                {"member": row["Contributor"], "commits": row["Commits"]}
                for row in activity["contributors"]
            ]
            st.altair_chart(dashboard_theme.hbar_chart(
                pd.DataFrame(commit_load), "commits", "member",
                chart_theme, color=ACCENT["blue"],
            ), width="stretch")
    pr_status = Counter(row.get("Status", "") for row in activity["pull_requests"])
    status_col, inventory_col = st.columns([1, 1.4])
    with status_col:
        section_header("Pull request outcomes", "نتائج Pull Requests", "⌥")
        if pr_status:
            st.altair_chart(dashboard_theme.donut_chart(
                pd.DataFrame({"status": list(pr_status), "total": list(pr_status.values())}),
                "status", "total", chart_theme,
                colors={"completed": ACCENT["green"], "active": ACCENT["blue"],
                        "abandoned": ACCENT["red"]},
            ), width="stretch")
    with inventory_col:
        _activity_table("Repository inventory", "قائمة المستودعات", activity["repositories"])
    _activity_table("Complete commit history", "كل تاريخ Commits", activity["commits"], 520)
    _activity_table("Complete push history", "كل تاريخ عمليات الرفع", activity["pushes"], 480)
    _activity_table("All pull requests", "كل Pull Requests", activity["pull_requests"], 520)
    _activity_table("All changed files", "كل الملفات المتغيرة", activity["changes"], 600)
    if not failures:
        return
    st.warning(tr(
        "Some repository API calls failed; all successful data is still shown below.",
        "فشلت بعض استدعاءات المستودعات؛ ما زالت كل البيانات الناجحة معروضة.",
    ))
    _activity_table("Collection failures", "أخطاء جمع البيانات", failures)


def render_repository_intelligence():
    st.header(tr("Repository Intelligence", "ذكاء المستودعات"))
    st.caption(tr(
        "All-history Azure Repos inventory, contributor activity, commits, pushes, PRs and changed files.",
        "كل تاريخ مستودعات Azure: المساهمون وCommits وعمليات الرفع وPull Requests والملفات المتغيرة.",
    ))
    _clear_repository_cache_when_requested()
    _fetch_repository_cache()
    activity = _repository_cache()
    if not activity:
        return
    filtered = _filtered_repository_activity(
        activity, _repository_filter_controls(activity)
    )
    _render_repository_kpis(filtered)
    _render_repository_tables(filtered, activity["failures"])

# ============================================================ RELEASES
def render_releases():
    st.header(tr("Releases", "الإصدارات"))
    # Release plans / target dates are not part of the Azure work-item pull.
    # This mirrors the workbook's "Releases" tab (release intelligence / manual input).
    release_rows = [{
        "Version": "No release-plan source in current Azure pull",
        "Platform": "", "Target Date": None, "Actual Date": None,
        "Status": "", "Owner": "", "Release Notes": "Connect Azure Pipelines/Delivery Plans or a Target-Date field to populate.",
    }]
    st.dataframe(localized_frame(pd.DataFrame(release_rows)), width="stretch", hide_index=True)
    st.info(tr(
        "Release dates are not present in the work-item pull. Add a Target Date field or connect Azure Pipelines.",
        "تواريخ الإصدارات غير موجودة في سحب عناصر العمل. أضف حقل Target Date أو اربط Azure Pipelines.",
    ))


# ============================================================ RAW DATA
def render_raw_data():
    st.header(tr("Raw Data — all Azure work items", "البيانات الخام — جميع عناصر Azure"))
    st.caption(tr(f"{len(items):,} items from the current {data_mode} source.", f"عدد {len(items):,} عنصر من مصدر البيانات الحالي."))
    all_items = items  # already loaded by the app
    df = pd.DataFrame([{
        "Work Item ID": i["id"], "Title": i["title"], "Work Item Type": i["type"],
        "State": i["state"], "State Category": i["state_category"],
        "Board Column": i["board_column"], "Board Column Done": i["board_column_done"],
        "Board Lane": i["board_lane"], "Assigned To": i["assignee"],
        "Iteration Path": i["sprint"], "Area Path": i["area"],
        "Story Points": i["sp"], "Priority": i["priority"],
        "Created Date": i["created"], "Changed Date": i["changed"],
        "Tags": "; ".join(i["tags"]) or "Untagged", "Parent ID": i["parent"], "Azure Link": i["url"],
    } for i in all_items])
    if not df.empty:
        st.dataframe(localized_frame(df), width="stretch", hide_index=True, height=520)
    else:
        st.info(tr(
            "No raw data loaded. Sync Azure DevOps or provide a workbook.",
            "لا توجد بيانات خام. قم بمزامنة Azure DevOps أو أضف ملف البيانات.",
        ))


# ---- dispatch
if page == "Executive Dashboard":
    render_executive()
elif page == "Sprint Summary":
    render_sprint_summary()
elif page == "Sprint Board":
    render_sprint_board()
elif page == "Tag Analysis":
    render_tag_analysis()
elif page == "Team Analysis":
    render_team_analysis()
elif page == "Area Analysis":
    render_area_analysis()
elif page == "Active Now":
    render_active_now()
elif page == "Risks & Aging":
    render_risks()
elif page == "Data Quality":
    render_data_quality()
elif page == "Repository Intelligence":
    render_repository_intelligence()
elif page == "Releases":
    render_releases()
elif page == "Raw Data":
    render_raw_data()