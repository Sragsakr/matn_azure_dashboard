"""Build a professional Azure DevOps delivery analytics workbook.

The Raw Data sheet is the source of truth. All analytical sheets are rebuilt
on every run, so no sample/manual value can be mistaken for Azure data.

Run after pulling Azure data:
    ./.venv/bin/python3 build_dashboard.py
"""

from __future__ import annotations

import datetime as dt
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl
from openpyxl.chart import BarChart, DoughnutChart, Reference
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

WORKBOOK = Path("Delivery_Manager_Dashboard.xlsx")
TODAY = dt.date.today()
DELIVERY_TYPES = ("Epic", "Feature", "User Story", "Task", "Bug")
DEV_TYPES = set(DELIVERY_TYPES)
# Fallbacks for older workbook caches; fresh pulls carry Azure state categories.
DONE_STATES = {"Closed", "Done", "Resolved", "Completed"}
ACTIVE_STATES = {"Active", "In Progress", "Committed"}
TERMINAL_CATEGORIES = {"Completed", "Removed"}
PROJECT_ROOT = "Hoteliana"

# --- Analysis thresholds (single source of truth for the health verdicts) ---
STALE_DAYS = 14          # open item older than this is "stale"/aging
VELOCITY_COLOR_HI = 0.70 # metric >= this -> green (on pace)
VELOCITY_COLOR_MID = 0.40# metric 0.40..0.69 -> amber (watch)
UNASSIGNED_ALERT = 25    # reorganize allocation when this many are ownerless
STALE_ALERT = 10         # urgent when at least this many open items are stale
# ----------------------------------------------------------------------------

# Visual system
NAVY = "17243B"
NAVY_2 = "223451"
BLUE = "2F75B5"
CYAN = "23A6D5"
GREEN = "25A66A"
AMBER = "F4B942"
RED = "E85D5D"
PURPLE = "7057D9"
WHITE = "FFFFFF"
INK = "1E293B"
LIGHT = "F5F7FB"
MUTED = "64748B"
LIGHT_GREEN = "E8F5EE"
LIGHT_AMBER = "FFF5DC"
LIGHT_RED = "FDECEC"
GRID = "D9E2EF"

THIN = Side(style="thin", color=GRID)


def parse_date(value):
    if value is None or value == "":
        return None
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    try:
        return dt.datetime.fromisoformat(str(value).replace("Z", "+00:00")).date()
    except ValueError:
        try:
            return dt.datetime.strptime(str(value)[:10], "%Y-%m-%d").date()
        except ValueError:
            return None


def leaf(path):
    if not path:
        return ""
    return str(path).replace("\\\\", "\\").split("\\")[-1]


def sprint_name(path):
    """Project-only iteration is Product Backlog; child path is a sprint."""
    if not path:
        return "Product Backlog"
    normalized = str(path).replace("\\\\", "\\")
    return normalized.split("\\")[-1] if "\\" in normalized else "Product Backlog"


def split_tags(value):
    if not value:
        return []
    return [tag.strip() for tag in str(value).split(";") if tag.strip()]


def state_category(state, category=None):
    if category:
        return category
    if state in DONE_STATES:
        return "Completed"
    if state in ACTIVE_STATES:
        return "InProgress"
    if str(state).lower() in {"removed", "deleted", "cut"}:
        return "Removed"
    return "Proposed"


def is_done(item):
    return item["state_category"] == "Completed"


def is_open(item):
    return item["state_category"] not in TERMINAL_CATEGORIES


def is_active(item):
    return item["state_category"] == "InProgress"


def read_raw(wb):
    ws = wb["Raw Data"]
    headers = {
        str(cell.value).strip(): idx
        for idx, cell in enumerate(ws[5])
        if cell.value is not None
    }

    required = {
        "Work Item ID", "Title", "Work Item Type", "State", "Assigned To",
        "Iteration Path", "Area Path", "Story Points", "Priority",
        "Created Date", "Changed Date", "Tags", "URL",
    }
    missing = required - headers.keys()
    if missing:
        raise ValueError(f"Raw Data is missing columns: {sorted(missing)}")

    def value(row, name, default=None):
        idx = headers.get(name)
        return row[idx] if idx is not None and idx < len(row) else default

    items = []
    for row in ws.iter_rows(min_row=6, values_only=True):
        if value(row, "Work Item ID") is None:
            continue
        iteration_path = value(row, "Iteration Path")
        created = parse_date(value(row, "Created Date"))
        changed = parse_date(value(row, "Changed Date"))
        closed = parse_date(value(row, "Closed Date"))
        state = value(row, "State") or "Unknown"
        item = {
            "id": value(row, "Work Item ID"),
            "title": value(row, "Title") or "",
            "type": value(row, "Work Item Type") or "Unknown",
            "state": state,
            "state_category": state_category(state, value(row, "State Category")),
            "board_column": value(row, "Board Column") or state,
            "board_lane": value(row, "Board Lane") or "Default",
            "assignee": value(row, "Assigned To") or "Unassigned",
            "iteration_path": iteration_path or "",
            "sprint": sprint_name(iteration_path),
            "area": leaf(value(row, "Area Path")) or PROJECT_ROOT,
            "sp": value(row, "Story Points"),
            "priority": value(row, "Priority"),
            "created": created,
            "changed": changed,
            "closed": closed,
            "tags": split_tags(value(row, "Tags")),
            "url": value(row, "URL") or "",
            "parent": value(row, "Parent ID"),
        }
        item["age"] = (TODAY - created).days if created else None
        # True Closed Date is preferred. Changed Date is not treated as a true
        # closure timestamp; it is only exposed separately for transparency.
        item["lead_time"] = (
            (closed - created).days if closed and created and is_done(item) else None
        )
        items.append(item)
    return items


def recreate_sheet(wb, name, index=None):
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name, index if index is not None else len(wb.sheetnames))
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.sheet_view.zoomScale = 90
    return ws


def title(ws, text, subtitle, end_col=12):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_col)
    ws["A1"] = text
    ws["A1"].font = Font(name="Aptos Display", size=22, bold=True, color=WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A1"].alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 38

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=end_col)
    ws["A2"] = subtitle
    ws["A2"].font = Font(name="Aptos", size=10, color="DCE6F2")
    ws["A2"].fill = PatternFill("solid", fgColor=NAVY_2)
    ws["A2"].alignment = Alignment(vertical="center")
    ws.row_dimensions[2].height = 24


def section(ws, row, text, start_col=1, end_col=12):
    ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
    cell = ws.cell(row, start_col, text)
    cell.font = Font(name="Aptos Display", size=12, bold=True, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=NAVY_2)
    cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[row].height = 23


def header_row(ws, row, headers, start_col=1):
    for offset, value in enumerate(headers):
        cell = ws.cell(row, start_col + offset, value)
        cell.font = Font(name="Aptos", size=10, bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=Side(style="medium", color=NAVY))
    ws.row_dimensions[row].height = 28


def card(ws, row, col, label, value, color=BLUE, number_format=None, width=2):
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + width - 1)
    ws.merge_cells(start_row=row + 1, start_column=col, end_row=row + 2, end_column=col + width - 1)
    label_cell = ws.cell(row, col, label)
    value_cell = ws.cell(row + 1, col, value)
    for rr in range(row, row + 3):
        for cc in range(col, col + width):
            c = ws.cell(rr, cc)
            c.fill = PatternFill("solid", fgColor=color)
            c.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    label_cell.font = Font(name="Aptos", size=9, bold=True, color=WHITE)
    label_cell.alignment = Alignment(horizontal="center", vertical="center")
    value_cell.font = Font(name="Aptos Display", size=20, bold=True, color=WHITE)
    value_cell.alignment = Alignment(horizontal="center", vertical="center")
    if number_format:
        value_cell.number_format = number_format
    ws.row_dimensions[row].height = 20
    ws.row_dimensions[row + 1].height = 26
    ws.row_dimensions[row + 2].height = 16


def style_table_body(ws, min_row, max_row, min_col, max_col):
    for row in range(min_row, max_row + 1):
        fill = PatternFill("solid", fgColor=WHITE if row % 2 else LIGHT)
        for col in range(min_col, max_col + 1):
            cell = ws.cell(row, col)
            cell.fill = fill
            cell.border = Border(bottom=THIN)
            cell.font = Font(name="Aptos", size=9, color=INK)
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def add_table(ws, ref, name):
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=False,
    )
    ws.add_table(table)


def set_widths(ws, widths):
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def item_metrics(items):
    total = len(items)
    done = sum(is_done(i) for i in items)
    active = sum(is_active(i) for i in items)
    open_count = sum(is_open(i) for i in items)
    unassigned = sum(
        i["type"] == "Task" and i["assignee"] == "Unassigned" for i in items
    )
    stale = sum(
        is_open(i) and i["age"] is not None and i["age"] >= STALE_DAYS
        for i in items
    )
    return {
        "total": total,
        "done": done,
        "active": active,
        "open": open_count,
        "unassigned": unassigned,
        "stale": stale,
        "done_pct": done / total if total else 0,
    }


def scope_metrics(items):
    """A User Story counts as Done when it is actually closed AND, when a
    hierarchy is available, every child Task is also closed. Falls back to the
    story's own state when no Parent ID is populated (data not yet upgraded)."""
    stories = [i for i in items if i["type"] == "User Story"]
    tasks = [i for i in items if i["type"] == "Task"]

    # Build child-task lookup when the enhanced pull has populated Parent ID.
    tasks_by_parent = defaultdict(list)
    hierarchy_available = False
    for task in tasks:
        if task["parent"] is not None:
            tasks_by_parent[task["parent"]].append(task)
            hierarchy_available = True

    done_stories = []
    for story in stories:
        if not is_done(story):
            continue
        child_tasks = tasks_by_parent.get(story["id"])
        if child_tasks and hierarchy_available:
            # Story is Done only if every linked child Task is also Done.
            if all(is_done(t) for t in child_tasks):
                done_stories.append(story)
        else:
            # No hierarchy for this story: trust its own state.
            done_stories.append(story)

    done_tasks = [i for i in tasks if is_done(i)]
    total_sp = sum(float(i["sp"] or 0) for i in stories)
    done_sp = sum(float(i["sp"] or 0) for i in done_stories)
    return {
        "stories": len(stories),
        "stories_done": len(done_stories),
        "scope_pct": len(done_stories) / len(stories) if stories else None,
        "tasks": len(tasks),
        "tasks_done": len(done_tasks),
        "task_pct": len(done_tasks) / len(tasks) if tasks else None,
        "total_sp": total_sp,
        "done_sp": done_sp,
        "velocity_pct": done_sp / total_sp if total_sp else None,
        "hierarchy_used": hierarchy_available,
    }


def type_progress(items):
    """Independent completion % per work type, plus genuine child→parent
    roll-up when Parent ID is present (Epic←Feature←User Story←Task).
    Without Parent ID it reports each type's own closed-state percentage
    and marks hierarchy as unavailable."""
    by_type = defaultdict(list)
    for item in items:
        by_type[item["type"]].append(item)

    result = {}
    for t in DELIVERY_TYPES:
        group = by_type.get(t, [])
        result[t] = {
            "total": len(group),
            "done": sum(1 for i in group if is_done(i)),
            "pct": None,
        }
    for t in result:
        total = result[t]["total"]
        result[t]["pct"] = (result[t]["done"] / total) if total else None

    hierarchy_used = any(item["parent"] is not None for item in items)

    if hierarchy_used:
        children = defaultdict(list)
        for item in items:
            if item["parent"] is not None:
                children[item["parent"]].append(item)

        def _done(vertex):
            # Built-in types are Done by closed state.
            return is_done(vertex)

        # Every vertex's effective Done = self-done AND all children effective-done
        # (child task/feature/story must all be complete for a parent to count Done).
        def effective_done(vertex, memo):
            if vertex["id"] in memo:
                return memo[vertex["id"]]
            res = _done(vertex)
            for child in children.get(vertex["id"], []):
                if not effective_done(child, memo):
                    res = False
                    break
            memo[vertex["id"]] = res
            return res

        memo = {}
        epic_done = 0
        feature_done = 0
        story_done = 0
        task_done = 0
        for item in items:
            if effective_done(item, memo):
                if item["type"] == "Epic":
                    epic_done += 1
                elif item["type"] == "Feature":
                    feature_done += 1
                elif item["type"] == "User Story":
                    story_done += 1
                elif item["type"] == "Task":
                    task_done += 1
        result["Task"]["done"] = task_done
        result["User Story"]["done"] = story_done
        result["Feature"]["done"] = feature_done
        result["Epic"]["done"] = epic_done
        for t in result:
            total = result[t]["total"]
            result[t]["pct"] = (result[t]["done"] / total) if total else None

    result["hierarchy_used"] = hierarchy_used
    return result


def value_color(value):
    """Traffic-light colour for a 0..1 metric (green>=HI, amber>=MID, else red)."""
    if value is None:
        return MUTED
    if value >= VELOCITY_COLOR_HI:
        return GREEN
    if value >= VELOCITY_COLOR_MID:
        return AMBER
    return RED


def state_fill(state):
    """Fill color for a work state (done=green, active=blue, blocked=red, else amber)."""
    if state in DONE_STATES:
        return GREEN
    if state in ACTIVE_STATES:
        return BLUE
    if state.lower() in {"blocked", "waiting", "on hold"}:
        return RED
    return AMBER


def ribbon_status(scope, task, unassigned, stale):
    """Overall health verdict from the four core signals using shared thresholds."""
    red_flags = 0
    if (scope or 0) < VELOCITY_COLOR_MID:
        red_flags += 1
    if (task or 0) < VELOCITY_COLOR_MID:
        red_flags += 1
    if unassigned and unassigned >= UNASSIGNED_ALERT:
        red_flags += 1
    if stale and stale >= STALE_ALERT:
        red_flags += 1
    if red_flags >= 3:
        return "CRITICAL", RED
    if red_flags >= 2:
        return "AT RISK", AMBER
    return "HEALTHY", GREEN


def delivery_action(items):
    """A one-line management action, choosing the highest-priority concern."""
    scope = scope_metrics(items)
    unassigned = sum(1 for i in items if i["assignee"] == "Unassigned")
    active = sum(is_active(i) for i in items)
    open_count = sum(is_open(i) for i in items)

    # The unassigned queue is the fastest, highest-impact lever; act on it
    # before anything else unless the sprint is already fully closed/done.
    if items and unassigned >= UNASSIGNED_ALERT and open_count:
        return f"{unassigned} work items are unassigned — allocate owners before sprint commitment"
    if items and scope["scope_pct"] == 0 and open_count:
        return "Scope stalled — no closed user stories; plan story closure"
    if any(i["type"] == "User Story" and i["sp"] is None for i in items):
        return "Add Story Points to unestimated stories for velocity"
    if sum(not i["tags"] for i in items):
        return "Add tags (Backend/Frontend/QA) for reliable analysis"
    if active == 0 and open_count:
        return "Nothing in progress — unblock or start assigned work"
    return "Delivery on track — keep closure rate steady"


def health_banner(ws, row, value, color, start_col=1, end_col=12):
    """A full-width status ribbon with a colored verdict."""
    ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
    cell = ws.cell(row, start_col, value)
    cell.font = Font(name="Aptos Display", size=16, bold=True, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=color)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 34
    return color


def build_executive(wb, dev):
    ws = recreate_sheet(wb, "Executive Dashboard", 0)
    title(
        ws,
        "HOTELIANA  •  DELIVERY CONTROL TOWER",
        f"Azure DevOps mirror  |  refreshed {TODAY:%d %b %Y}  |  scope completion uses User Stories only",
        12,
    )
    sprint_items = [i for i in dev if i["sprint"] != "Product Backlog"]
    backlog = [i for i in dev if i["sprint"] == "Product Backlog"]
    all_m = item_metrics(dev)
    scope = scope_metrics(sprint_items)
    prog = type_progress(dev)

    card(ws, 4, 1, "SPRINT SCOPE DONE", scope["scope_pct"] or 0, GREEN, "0%")
    card(ws, 4, 3, "TASK COMPLETION", scope["task_pct"] or 0, BLUE, "0%")
    card(ws, 4, 5, "SP VELOCITY", scope["velocity_pct"] or 0, PURPLE, "0%")
    card(ws, 4, 7, "ACTIVE NOW", all_m["active"], CYAN)
    card(ws, 4, 9, "UNASSIGNED", all_m["unassigned"], RED)
    card(ws, 4, 11, "NO-SPRINT BACKLOG", len(backlog), AMBER)

    verdict, color = ribbon_status(
        scope["scope_pct"], scope["task_pct"], all_m["unassigned"], all_m["stale"]
    )
    health_banner(ws, 7, f"{verdict}  —  {delivery_action(dev)}", color)

    section(ws, 9, "DELIVERY BREAKDOWN — independent % per type", 1, 6)
    hierarchy_note = "(rolled up child→parent)" if prog["hierarchy_used"] else "(own-state — pull enhanced data for roll-up)"
    header_row(ws, 10, ["Work Type", "Total", "Done", "Completion %", hierarchy_note], 1)
    tr = 11
    for t in DELIVERY_TYPES:
        meta = prog[t]
        ws.cell(tr, 1, t)
        ws.cell(tr, 2, meta["total"])
        ws.cell(tr, 3, meta["done"])
        ws.cell(tr, 4, meta["pct"]).number_format = "0%"
        ws.cell(tr, 4).font = Font(bold=True, color=value_color(meta["pct"]))
        tr += 1
    style_table_body(ws, 11, tr - 1, 1, 5)

    section(ws, 9, "DELIVERY HEALTH — exact Azure states", 7, 12)
    state_counts = Counter(i["state"] for i in dev)
    header_row(ws, 10, ["State", "Items", "% of dev scope"], 7)
    row = 11
    for state, count in state_counts.most_common():
        ws.cell(row, 7, state)
        ws.cell(row, 8, count)
        ws.cell(row, 9, count / len(dev) if dev else 0).number_format = "0%"
        row += 1
    style_table_body(ws, 11, row - 1, 7, 9)

    product_section_row = max(tr, row) + 1
    product_header_row = product_section_row + 1
    section(ws, product_section_row, "PRODUCT AREAS", 7, 12)
    header_row(ws, product_header_row, ["Area", "Stories", "Done", "Scope %", "Tasks", "Task %"], 7)
    area_groups = defaultdict(list)
    for item in dev:
        area_groups[item["area"]].append(item)
    ar = product_header_row + 1
    for area, items in sorted(area_groups.items(), key=lambda x: -len(x[1])):
        m = scope_metrics(items)
        values = [area, m["stories"], m["stories_done"], m["scope_pct"], m["tasks"], m["task_pct"]]
        for col, value in enumerate(values, 7):
            ws.cell(ar, col, value)
        ws.cell(ar, 10).number_format = "0%"
        ws.cell(ar, 12).number_format = "0%"
        ar += 1
    style_table_body(ws, product_header_row + 1, ar - 1, 7, 12)

    alert_row_offset = ar
    section(ws, alert_row_offset + 1, "MANAGEMENT ALERTS", 1, 6)
    alerts = [
        ("Unassigned work", all_m["unassigned"], "Assign an owner before sprint commitment"),
        (f"Open ≥ {STALE_DAYS} days", all_m["stale"], "Review items with no closure for aging risk"),
        ("Product Backlog (no sprint)", len(backlog), "These items are not committed to any sprint"),
        ("Items without tags", sum(not i["tags"] for i in dev), "Improve classification for reliable analysis"),
        ("Stories without Story Points", sum(i["type"] == "User Story" and i["sp"] is None for i in dev), "Velocity is incomplete without estimation"),
    ]
    header_row(ws, alert_row_offset + 2, ["Signal", "Count", "Management action"], 1)
    for idx, values in enumerate(alerts, alert_row_offset + 3):
        for col, value in enumerate(values, 1):
            ws.cell(idx, col, value)
    style_table_body(ws, alert_row_offset + 3, alert_row_offset + 3 + len(alerts) - 1, 1, 3)

    def_row_offset = alert_row_offset + 3 + len(alerts)
    section(ws, def_row_offset + 1, "MEASUREMENT DEFINITIONS", 7, 12)
    definitions = [
        ("Sprint Scope Done", "Closed User Stories ÷ all User Stories in a sprint"),
        ("Task Completion", "Closed Tasks ÷ all Tasks in a sprint"),
        ("SP Velocity", "Story Points on Closed User Stories ÷ committed Story Points"),
        ("All-item closure", "Diagnostic only; never used as project scope completion"),
        ("Product Backlog", f"Iteration Path equals project root '{PROJECT_ROOT}' (no sprint child)"),
    ]
    header_row(ws, def_row_offset + 2, ["Metric", "Definition"], 7)
    for idx, values in enumerate(definitions, def_row_offset + 3):
        ws.cell(idx, 7, values[0])
        ws.cell(idx, 8, values[1])
        ws.merge_cells(start_row=idx, start_column=8, end_row=idx, end_column=12)
    style_table_body(ws, def_row_offset + 3, def_row_offset + 3 + len(definitions) - 1, 7, 12)

    # Charts read visible sections above for auditability.
    chart = DoughnutChart()
    chart.title = "Work items by Azure state"
    # state table now lives at columns 7-9 (Items in col 8, State in col 7)
    first_data = 11
    last_data = row - 1
    chart.add_data(Reference(ws, min_col=8, min_row=first_data, max_row=last_data), titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=7, min_row=first_data, max_row=last_data))
    chart.height = 7.3
    chart.width = 12
    chart.legend.position = "r"
    ws.add_chart(chart, "A27")

    bar = BarChart()
    bar.type = "bar"
    bar.style = 10
    bar.title = "Area scope — User Stories"
    bar.y_axis.title = "Area"
    bar.x_axis.title = "Stories"
    bar.add_data(
        Reference(ws, min_col=8, min_row=product_header_row, max_row=ar - 1),
        titles_from_data=True,
    )
    bar.set_categories(
        Reference(ws, min_col=7, min_row=product_header_row + 1, max_row=ar - 1)
    )
    bar.height = 7.3
    bar.width = 12
    ws.add_chart(bar, "G27")

    set_widths(ws, {"A": 22, "B": 14, "C": 24, "D": 14, "E": 14, "F": 3,
                    "G": 26, "H": 22, "I": 12, "J": 13, "K": 13, "L": 13})
    ws.freeze_panes = "A3"


def build_sprint_summary(wb, dev):
    ws = recreate_sheet(wb, "Sprint Summary", 1)
    title(ws, "SPRINT PORTFOLIO", "One row per real Azure iteration; Product Backlog is explicitly separated", 15)
    headers = [
        "Iteration", "Total Dev Items", "User Stories", "Stories Done", "Scope Done %",
        "Tasks", "Tasks Done", "Task Done %", "Committed SP", "Done SP", "SP Velocity %",
        "Active", "Unassigned", f"Open ≥{STALE_DAYS}d", "Exact Iteration Meaning",
    ]
    header_row(ws, 4, headers)
    groups = defaultdict(list)
    for item in dev:
        groups[item["sprint"]].append(item)
    ordered = sorted(groups, key=lambda value: (value == "Product Backlog", value))
    row = 5
    for sprint in ordered:
        items = groups[sprint]
        sm = scope_metrics(items)
        im = item_metrics(items)
        values = [
            sprint, len(items), sm["stories"], sm["stories_done"], sm["scope_pct"],
            sm["tasks"], sm["tasks_done"], sm["task_pct"], sm["total_sp"], sm["done_sp"],
            sm["velocity_pct"], im["active"], im["unassigned"], im["stale"],
            "Committed iteration" if sprint != "Product Backlog" else "No sprint assigned",
        ]
        for col, value in enumerate(values, 1):
            ws.cell(row, col, value)
        for col, metric in ((5, sm["scope_pct"]), (8, sm["task_pct"]), (11, sm["velocity_pct"])):
            ws.cell(row, col).number_format = "0%"
            ws.cell(row, col).font = Font(bold=True, color=value_color(metric))
        row += 1
    style_table_body(ws, 5, row - 1, 1, len(headers))
    add_table(ws, f"A4:O{row - 1}", "SprintSummaryTable")
    ws.conditional_formatting.add(f"E5:E{row - 1}", DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color=GREEN))
    ws.conditional_formatting.add(f"H5:H{row - 1}", DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color=BLUE))
    ws.conditional_formatting.add(f"K5:K{row - 1}", DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color=PURPLE))
    ws.auto_filter.ref = f"A4:O{row - 1}"
    ws.freeze_panes = "A5"
    set_widths(ws, {"A": 22, "B": 14, "C": 13, "D": 13, "E": 14, "F": 10, "G": 12,
                    "H": 13, "I": 14, "J": 10, "K": 14, "L": 10, "M": 12, "N": 12, "O": 22})


def build_sprint_board(wb, dev):
    ws = recreate_sheet(wb, "Sprint Board", 2)
    title(ws, "SPRINT WORK ITEM EXPLORER", "Every dev work item grouped by real Azure iteration and assignee", 15)
    headers = [
        "Iteration", "Work Item ID", "Title", "Type", "State", "Assignee", "Area",
        "Tags", "Story Points", "Priority", "Created", "Changed", "Age (days)", "Parent ID", "Azure Link",
    ]
    header_row(ws, 4, headers)
    ordered = sorted(dev, key=lambda i: (i["sprint"] == "Product Backlog", i["sprint"], i["assignee"], i["type"], i["id"]))
    row = 5
    for item in ordered:
        values = [
            item["sprint"], item["id"], item["title"], item["type"], item["state"],
            item["assignee"], item["area"], "; ".join(item["tags"]) or "Untagged",
            item["sp"], item["priority"], item["created"], item["changed"], item["age"],
            item["parent"], "Open in Azure",
        ]
        for col, value in enumerate(values, 1):
            ws.cell(row, col, value)
        ws.cell(row, 5).fill = PatternFill("solid", fgColor=state_fill(item["state"]))
        ws.cell(row, 5).font = Font(bold=True, color=WHITE)
        if item["url"]:
            ws.cell(row, 15).hyperlink = item["url"]
            ws.cell(row, 15).style = "Hyperlink"
        row += 1
    style_table_body(ws, 5, row - 1, 1, len(headers))
    add_table(ws, f"A4:O{row - 1}", "SprintWorkItemsTable")
    ws.auto_filter.ref = f"A4:O{row - 1}"
    ws.freeze_panes = "C5"
    ws.conditional_formatting.add(f"M5:M{row - 1}", ColorScaleRule(start_type="min", start_color=LIGHT_GREEN, mid_type="percentile", mid_value=50, mid_color=LIGHT_AMBER, end_type="max", end_color=LIGHT_RED))
    set_widths(ws, {"A": 20, "B": 13, "C": 48, "D": 14, "E": 14, "F": 24, "G": 25,
                    "H": 28, "I": 12, "J": 10, "K": 13, "L": 13, "M": 12, "N": 12, "O": 16})


def build_tag_analysis(wb, dev):
    ws = recreate_sheet(wb, "Tag Analysis", 3)
    title(ws, "TAG ANALYSIS", "Multi-tag items are counted once under each tag; Untagged is shown explicitly", 12)
    groups = defaultdict(list)
    for item in dev:
        tags = item["tags"] or ["Untagged"]
        for tag in tags:
            groups[tag].append(item)

    headers = ["Tag", "Items", "Stories", "Stories Done", "Scope Done %", "Tasks", "Tasks Done", "Task Done %", "Active", "Unassigned", f"Open ≥{STALE_DAYS}d", "Areas", "Management focus"]
    header_row(ws, 4, headers)
    row = 5
    for tag, items in sorted(groups.items(), key=lambda x: (-len(x[1]), x[0])):
        sm = scope_metrics(items)
        im = item_metrics(items)
        values = [
            tag, len(items), sm["stories"], sm["stories_done"], sm["scope_pct"],
            sm["tasks"], sm["tasks_done"], sm["task_pct"], im["active"],
            im["unassigned"], im["stale"], ", ".join(sorted({i["area"] for i in items})),
            delivery_action(items),
        ]
        for col, value in enumerate(values, 1):
            ws.cell(row, col, value)
        for col, metric in ((5, sm["scope_pct"]), (8, sm["task_pct"])):
            ws.cell(row, col).number_format = "0%"
            ws.cell(row, col).font = Font(bold=True, color=value_color(metric))
        row += 1
    style_table_body(ws, 5, row - 1, 1, len(headers))
    add_table(ws, f"A4:M{row - 1}", "TagAnalysisTable")
    ws.conditional_formatting.add(f"E5:E{row - 1}", DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color=GREEN))
    ws.conditional_formatting.add(f"H5:H{row - 1}", DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color=BLUE))
    ws.freeze_panes = "A5"
    set_widths(ws, {"A": 22, "B": 10, "C": 11, "D": 13, "E": 14, "F": 10, "G": 12,
                    "H": 13, "I": 10, "J": 12, "K": 12, "L": 40, "M": 40})


def build_team_analysis(wb, dev):
    ws = recreate_sheet(wb, "Team Analysis", 4)
    title(ws, "TEAM DELIVERY ANALYSIS", "Task-centric view — a User Story is rarely owned by one person; deliverable progress is credited through the Tasks a member made Done", 14)
    headers = [
        "Assignee", "Tasks", "Tasks Done", "Task Done %", "Active", "Open",
        f"Open ≥{STALE_DAYS}d", "Stories Involved", "Stories Fully Done", "Areas", "Tags", "Exact Open Work",
    ]
    header_row(ws, 4, headers)
    groups = defaultdict(list)
    for item in dev:
        groups[item["assignee"]].append(item)

    # Story→assignees: credit every assignee who owns at least one child Task of a story.
    story_by_id = {item["id"]: item for item in dev if item["type"] == "User Story"}
    assignee_of_story = defaultdict(set)   # story_id -> set(assignees)
    for task in dev:
        if task["type"] == "Task" and task["parent"] in story_by_id:
            assignee_of_story[task["parent"]].add(task["assignee"])

    row = 5
    for assignee, items in sorted(groups.items(), key=lambda x: (-sum(1 for i in x[1] if i["type"] == "Task"), x[0])):
        tasks = [i for i in items if i["type"] == "Task"]
        done_tasks = sum(is_done(t) for t in tasks)
        im = item_metrics(items)
        # stories this assignee is part of (via a child task), and those fully done
        involved_stories = {i["id"] for i in items if i["type"] == "User Story"}
        for story_id, assignees in assignee_of_story.items():
            if assignee in assignees:
                involved_stories.add(story_id)
        fully_done_stories = {sid for sid in involved_stories
                              if story_by_id.get(sid) and is_done(story_by_id[sid])}
        current = [i for i in tasks if is_open(i)]
        work = " | ".join(f"#{i['id']} {i['title']} [{i['state']}]" for i in current)
        values = [
            assignee, len(tasks), done_tasks, done_tasks / len(tasks) if tasks else None,
            im["active"], im["open"], im["stale"], len(involved_stories),
            len(fully_done_stories),
            ", ".join(sorted({i["area"] for i in items})),
            ", ".join(sorted({t for i in items for t in i["tags"]})) or "Untagged",
            work or "No open work",
        ]
        for col, value in enumerate(values, 1):
            ws.cell(row, col, value)
        if tasks:
            ws.cell(row, 4).number_format = "0%"
            ws.cell(row, 4).font = Font(bold=True, color=value_color(done_tasks / len(tasks)))
        # highlight when a member is credited but a story isn't done
        if fully_done_stories and involved_stories - fully_done_stories:
            ws.cell(row, 5).font = Font(bold=True, color=AMBER)
        row += 1
    style_table_body(ws, 5, row - 1, 1, len(headers))
    add_table(ws, f"A4:L{row - 1}", "TeamAnalysisTable")
    ws.conditional_formatting.add(f"D5:D{row - 1}", DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color=BLUE))
    ws.freeze_panes = "B5"
    set_widths(ws, {"A": 24, "B": 9, "C": 12, "D": 13, "E": 9, "F": 9, "G": 11,
                    "H": 16, "I": 15, "J": 30, "K": 26, "L": 70})


def build_area_analysis(wb, dev):
    ws = recreate_sheet(wb, "Area Analysis", 5)
    title(ws, "PRODUCT AREA ANALYSIS", "Scope and execution split across Azure Area Paths", 13)
    headers = ["Area", "Total", "Stories", "Stories Done", "Scope Done %", "Tasks", "Tasks Done", "Task Done %", "SP", "Done SP", "Active", "Unassigned", f"Open ≥{STALE_DAYS}d", "Management focus"]
    header_row(ws, 4, headers)
    groups = defaultdict(list)
    for item in dev:
        groups[item["area"]].append(item)
    row = 5
    for area, items in sorted(groups.items(), key=lambda x: -len(x[1])):
        sm = scope_metrics(items)
        im = item_metrics(items)
        values = [area, len(items), sm["stories"], sm["stories_done"], sm["scope_pct"], sm["tasks"], sm["tasks_done"], sm["task_pct"], sm["total_sp"], sm["done_sp"], im["active"], im["unassigned"], im["stale"], delivery_action(items)]
        for col, value in enumerate(values, 1):
            ws.cell(row, col, value)
        for col, metric in ((5, sm["scope_pct"]), (8, sm["task_pct"])):
            ws.cell(row, col).number_format = "0%"
            ws.cell(row, col).font = Font(bold=True, color=value_color(metric))
        row += 1
    style_table_body(ws, 5, row - 1, 1, len(headers))
    add_table(ws, f"A4:N{row - 1}", "AreaAnalysisTable")
    ws.freeze_panes = "A5"
    set_widths(ws, {"A": 28, "B": 10, "C": 10, "D": 13, "E": 14, "F": 10, "G": 12,
                    "H": 13, "I": 10, "J": 10, "K": 10, "L": 12, "M": 12, "N": 40})


def build_active_board(wb, dev):
    """One screen of work the team should be touching now (open items), with
    risk colouring and a one-line action per owner/area."""
    ws = recreate_sheet(wb, "Active Now", 6)
    title(ws, "ACTIVE & OPEN WORK — TOUCH THIS NOW", "Everything open, risk-ranked; colour signals what to unblock first", 13)
    headers = ["Priority", "ID", "Title", "Type", "State", "Assignee", "Sprint", "Area", "Tags", "Age (days)", "Azure Link", "Why now"]
    header_row(ws, 4, headers)
    open_items = [i for i in dev if is_open(i)]
    ordered = sorted(open_items, key=lambda i: (i["assignee"] == "Unassigned", not is_active(i), -(i["age"] or -1), i["id"]))
    row = 5
    for item in ordered:
        priority = "High"
        if item["assignee"] == "Unassigned":
            priority = "Critical"
        elif is_active(item):
            priority = "Doing"
        elif item["age"] is not None and item["age"] >= STALE_DAYS:
            priority = "Aging"
        elif item["type"] == "User Story":
            priority = "Scope"
        why = {
            "Critical": "No owner — assign first",
            "Doing": "In progress — keep moving",
            "Aging": f"Open {item['age']}d — unblock or de-scope",
            "Scope": "User Story — scope requires it",
        }.get(priority, "Assigned — schedule it")
        color = {"Critical": RED, "Doing": BLUE, "Aging": AMBER}.get(priority, MUTED)
        values = [
            priority, item["id"], item["title"], item["type"], item["state"],
            item["assignee"], item["sprint"], item["area"], "; ".join(item["tags"]) or "Untagged",
            item["age"], "Open in Azure", why,
        ]
        for col, value in enumerate(values, 1):
            ws.cell(row, col, value)
        ws.cell(row, 1).font = Font(bold=True, color=WHITE)
        ws.cell(row, 1).fill = PatternFill("solid", fgColor=color)
        if item["url"]:
            ws.cell(row, 11).hyperlink = item["url"]
            ws.cell(row, 11).style = "Hyperlink"
        row += 1
    style_table_body(ws, 5, row - 1, 1, len(headers))
    add_table(ws, f"A4:L{row - 1}", "ActiveNowTable")
    ws.auto_filter.ref = f"A4:L{row - 1}"
    ws.freeze_panes = "B5"
    ws.conditional_formatting.add(f"J5:J{row - 1}", ColorScaleRule(start_type="min", start_color=LIGHT_GREEN, mid_type="percentile", mid_value=50, mid_color=LIGHT_AMBER, end_type="max", end_color=LIGHT_RED))
    set_widths(ws, {"A": 12, "B": 12, "C": 46, "D": 13, "E": 13, "F": 24, "G": 18,
                    "H": 24, "I": 22, "J": 12, "K": 14, "L": 34})


def build_risks(wb, dev):
    ws = recreate_sheet(wb, "Risks & Aging", 6)
    title(ws, "DELIVERY RISKS & AGING", "Open work ranked by age; age is days since Created Date, not time actively worked", 13)
    headers = ["Risk", "Age", "ID", "Title", "Type", "State", "Assignee", "Sprint", "Area", "Tags", "Created", "Changed", "Azure Link"]
    header_row(ws, 4, headers)
    open_items = [i for i in dev if is_open(i)]
    ordered = sorted(open_items, key=lambda i: (i["assignee"] != "Unassigned", -(i["age"] or -1), i["id"]))
    row = 5
    for item in ordered:
        risks = []
        if item["assignee"] == "Unassigned":
            risks.append("Unassigned")
        if item["sprint"] == "Product Backlog":
            risks.append("No sprint")
        if item["age"] is not None and item["age"] >= STALE_DAYS:
            risks.append(f"Age ≥{STALE_DAYS}d")
        if item["type"] == "User Story" and item["sp"] is None:
            risks.append("No SP")
        values = [", ".join(risks) or "Monitor", item["age"], item["id"], item["title"], item["type"], item["state"], item["assignee"], item["sprint"], item["area"], "; ".join(item["tags"]) or "Untagged", item["created"], item["changed"], "Open in Azure"]
        for col, value in enumerate(values, 1):
            ws.cell(row, col, value)
        if item["url"]:
            ws.cell(row, 13).hyperlink = item["url"]
            ws.cell(row, 13).style = "Hyperlink"
        row += 1
    style_table_body(ws, 5, row - 1, 1, len(headers))
    add_table(ws, f"A4:M{row - 1}", "RisksAgingTable")
    ws.conditional_formatting.add(f"B5:B{row - 1}", ColorScaleRule(start_type="min", start_color=LIGHT_GREEN, mid_type="percentile", mid_value=50, mid_color=LIGHT_AMBER, end_type="max", end_color=LIGHT_RED))
    ws.freeze_panes = "D5"
    set_widths(ws, {"A": 24, "B": 9, "C": 11, "D": 48, "E": 13, "F": 13, "G": 24,
                    "H": 20, "I": 26, "J": 24, "K": 13, "L": 13, "M": 16})


def build_data_quality(wb, all_items, dev):
    ws = recreate_sheet(wb, "Data Quality", 7)
    title(ws, "DATA QUALITY & TRUST", "What the workbook can prove from Azure — and what it refuses to invent", 10)
    section(ws, 4, "SOURCE COVERAGE", 1, 10)
    rows = [
        ("All Azure work items imported", len(all_items), "Exact"),
        ("Dev scope: Epic + Feature + User Story + Task", len(dev), "Exact"),
        ("Test artifacts retained in Raw Data", len(all_items) - len(dev), "Exact; excluded from delivery scope"),
        ("Items in real sprint paths", sum(i["sprint"] != "Product Backlog" for i in dev), "Exact"),
        ("Items at project root / no sprint", sum(i["sprint"] == "Product Backlog" for i in dev), "Exact"),
        ("Items without assignee", sum(i["assignee"] == "Unassigned" for i in dev), "Exact"),
        ("Items without tags", sum(not i["tags"] for i in dev), "Exact"),
        ("User Stories without Story Points", sum(i["type"] == "User Story" and i["sp"] is None for i in dev), "Exact"),
        ("Items with true Closed Date", sum(i["closed"] is not None for i in dev), "Requires enhanced Azure pull"),
        ("Items with Parent ID", sum(i["parent"] is not None for i in dev), "Requires enhanced Azure pull"),
    ]
    header_row(ws, 5, ["Check", "Count", "Interpretation"])
    for row, values in enumerate(rows, 6):
        for col, value in enumerate(values, 1):
            ws.cell(row, col, value)
    style_table_body(ws, 6, 5 + len(rows), 1, 3)

    section(ws, 18, "TRUTH RULES", 1, 10)
    truths = [
        ("No invented release dates", "Release dates are not present in the current Azure pull; the old sample Release rows were removed."),
        ("No fake Done Date", "Changed Date is not the same as Closed Date. Lead-time reporting stays unavailable until Closed Date is pulled."),
        ("No hierarchy double-counting", "Project/sprint scope completion uses User Stories, not Epic+Feature+Story+Task combined."),
        ("Tag counts overlap", "An item tagged Backend; Frontend; QA contributes once to every relevant tag."),
        ("Product Backlog is not a sprint", f"Iteration Path '{PROJECT_ROOT}' has no sprint child and is reported separately."),
    ]
    header_row(ws, 19, ["Rule", "Meaning"])
    for row, values in enumerate(truths, 20):
        ws.cell(row, 1, values[0])
        ws.cell(row, 2, values[1])
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=10)
    style_table_body(ws, 20, 24, 1, 10)
    set_widths(ws, {"A": 34, "B": 20, "C": 42, "D": 14, "E": 14, "F": 14, "G": 14, "H": 14, "I": 14, "J": 14})


def build_releases_truth(wb):
    ws = recreate_sheet(wb, "Releases", 8)
    title(ws, "RELEASE INTELLIGENCE", "No release-plan source is present in the current Azure pull", 8)
    section(ws, 4, "CURRENT STATUS", 1, 8)
    ws["A6"] = "No Azure-backed release dates are available yet."
    ws["A6"].font = Font(name="Aptos Display", size=16, bold=True, color=RED)
    ws.merge_cells("A6:H6")
    ws["A8"] = "Why"
    ws["B8"] = "The work-item pull includes iteration, state and dates, but not Azure Pipelines releases, Delivery Plans or target dates."
    ws["A9"] = "Truth policy"
    ws["B9"] = "The previous sample release rows were removed so this workbook cannot be mistaken for real Azure data."
    ws["A10"] = "Next integration"
    ws["B10"] = "Connect Azure Pipelines/Delivery Plans or add a dedicated Azure field for Target Date, then this sheet can be generated automatically."
    for row in range(8, 11):
        ws.cell(row, 1).font = Font(bold=True, color=NAVY)
        ws.cell(row, 2).alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
        ws.row_dimensions[row].height = 34
    set_widths(ws, {"A": 20, "B": 22, "C": 14, "D": 14, "E": 14, "F": 14, "G": 14, "H": 14})


def style_raw_data(wb):
    ws = wb["Raw Data"]
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A6"
    ws.auto_filter.ref = f"A5:{get_column_letter(ws.max_column)}{ws.max_row}"
    for cell in ws[5]:
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(bold=True, color=WHITE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[5].height = 30
    widths = [13, 52, 16, 14, 24, 28, 28, 12, 10, 13, 13, 28, 20, 13, 13]
    for idx in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(idx)].width = widths[idx - 1] if idx <= len(widths) else 16


def remove_legacy_sheets(wb):
    # These sheets were generated from older ambiguous logic. The new sheets
    # replace them with explicit scope/task/velocity definitions.
    for name in ("Dashboard", "Sprints", "Projects", "Team"):
        if name in wb.sheetnames:
            del wb[name]


def main():
    if not WORKBOOK.exists():
        raise FileNotFoundError(WORKBOOK)
    wb = openpyxl.load_workbook(WORKBOOK)
    all_items = read_raw(wb)
    dev = [item for item in all_items if item["type"] in DEV_TYPES]
    if not dev:
        raise ValueError("No Epic/Feature/User Story/Task items found in Raw Data")

    remove_legacy_sheets(wb)
    build_executive(wb, dev)
    build_sprint_summary(wb, dev)
    build_sprint_board(wb, dev)
    build_tag_analysis(wb, dev)
    build_team_analysis(wb, dev)
    build_area_analysis(wb, dev)
    build_active_board(wb, dev)
    build_risks(wb, dev)
    build_data_quality(wb, all_items, dev)
    build_releases_truth(wb)
    style_raw_data(wb)

    # Define the canonical tab order once, so scattered builder indices can
    # never leave ambiguous sheet ordering.
    order = [
        "Executive Dashboard", "Sprint Summary", "Sprint Board", "Tag Analysis",
        "Team Analysis", "Area Analysis", "Active Now", "Risks & Aging",
        "Data Quality", "Releases",
    ]
    target_index = 0
    for sheet_name in order:
        if sheet_name in wb.sheetnames:
            wb.move_sheet(sheet_name, offset=target_index - wb.sheetnames.index(sheet_name))
            target_index += 1

    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"
    wb.save(WORKBOOK)

    sprint_counts = Counter(i["sprint"] for i in dev)
    print(f"Saved professional Azure mirror: {WORKBOOK}")
    print(f"All work items: {len(all_items)} | Dev scope: {len(dev)}")
    print("Iterations:", dict(sprint_counts))
    print("Tags:", dict(Counter(t for i in dev for t in (i["tags"] or ["Untagged"]))))


if __name__ == "__main__":
    main()
