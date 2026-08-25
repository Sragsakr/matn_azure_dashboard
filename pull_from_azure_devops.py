"""
pull_from_azure_devops.py
--------------------------
Pulls work items from Azure DevOps (Boards) via the REST API and writes them
into the "Raw Data" sheet of Delivery_Manager_Dashboard.xlsx, without touching
any other sheet, formula, or formatting in the workbook.

Run this on YOUR machine (not in a shared environment) since it needs your
Azure DevOps Personal Access Token (PAT).

--------------------------------------------------------------------------
1) SETUP (one time)
--------------------------------------------------------------------------
    pip install requests openpyxl

Create a PAT in Azure DevOps:
    Azure DevOps > User settings (top right) > Personal access tokens > New Token
    Scope: "Work Items" -> Read (read-only is enough, and safest)

NEVER hardcode the PAT in this file or commit it to git. Set it as an
environment variable instead:

    # macOS/Linux
    export AZDO_PAT="your-token-here"

    # Windows (PowerShell)
    $env:AZDO_PAT = "your-token-here"

--------------------------------------------------------------------------
2) CONFIGURE
--------------------------------------------------------------------------
Fill in the four values below (org, project, team/iteration, workbook path).
"""

import os
import sys
import base64
import datetime
from urllib.parse import quote

import requests
import openpyxl

# ---- EDIT THESE ----------------------------------------------------------
ORG_NAME = "matnsolutions"                 # from https://dev.azure.com/your-org
PROJECT_NAME = "Hoteliana"         # exact project name in Azure DevOps
ITERATION_PATH = ""                   # e.g. "YourProject\\Sprint 14" — leave "" to pull ALL work items in the project
WORKBOOK_PATH = "Delivery_Manager_Dashboard.xlsx"
API_VERSION = "7.1"
# ----------------------------------------------------------------------------

PAT = os.environ.get("AZDO_PAT")
if not PAT:
    sys.exit(
        "ERROR: Set the AZDO_PAT environment variable first "
        "(see the setup instructions at the top of this script)."
    )

AUTH_HEADER = {
    "Authorization": "Basic "
    + base64.b64encode(f":{PAT}".encode()).decode(),
    "Content-Type": "application/json",
}

BASE_URL = f"https://dev.azure.com/{ORG_NAME}/{PROJECT_NAME}/_apis"


def run_wiql():
    """Get the list of work item IDs matching our filter via WIQL."""
    if ITERATION_PATH:
        where_clause = f"[System.IterationPath] = '{ITERATION_PATH}'"
    else:
        where_clause = f"[System.TeamProject] = '{PROJECT_NAME}'"

    query = {
        "query": (
            "SELECT [System.Id] FROM WorkItems "
            f"WHERE {where_clause} "
            "AND [System.WorkItemType] <> '' "
            "ORDER BY [System.ChangedDate] DESC"
        )
    }
    url = f"{BASE_URL}/wit/wiql?api-version={API_VERSION}"
    resp = requests.post(url, headers=AUTH_HEADER, json=query, timeout=30)
    resp.raise_for_status()
    data = resp.json()
    return [item["id"] for item in data.get("workItems", [])]


def get_work_items_batch(ids):
    """Fetch full field data for up to 200 work item IDs at a time."""
    fields = [
        "System.Id", "System.Title", "System.WorkItemType", "System.State",
        "System.AssignedTo", "System.IterationPath", "System.AreaPath",
        "System.Parent", "Microsoft.VSTS.Scheduling.StoryPoints",
        "Microsoft.VSTS.Common.Priority", "System.CreatedDate",
        "System.ChangedDate", "Microsoft.VSTS.Common.StateChangeDate",
        "Microsoft.VSTS.Common.ClosedDate", "System.Tags",
        "System.BoardColumn", "System.BoardColumnDone", "System.BoardLane",
    ]
    all_items = []
    for i in range(0, len(ids), 200):
        chunk = ids[i:i + 200]
        url = f"{BASE_URL}/wit/workitemsbatch?api-version={API_VERSION}"
        body = {"ids": chunk, "fields": fields}
        resp = requests.post(url, headers=AUTH_HEADER, json=body, timeout=30)
        resp.raise_for_status()
        all_items.extend(resp.json().get("value", []))
    return all_items


def enrich_state_categories(items):
    """Attach Azure's canonical category for every custom state."""
    work_types = {
        field(item, "System.WorkItemType", None)
        for item in items
        if field(item, "System.WorkItemType", None)
    }
    categories = {}
    for work_type in work_types:
        url = (
            f"{BASE_URL}/wit/workitemtypes/{quote(work_type, safe='')}/states"
            f"?api-version={API_VERSION}"
        )
        try:
            response = requests.get(url, headers=AUTH_HEADER, timeout=15)
            response.raise_for_status()
        except requests.RequestException as exc:
            print(f"WARNING: Could not read state metadata for {work_type}: {exc}")
            continue
        for state in response.json().get("value", []):
            categories[(work_type, state.get("name"))] = state.get("category")

    for item in items:
        item["_state_category"] = categories.get((
            field(item, "System.WorkItemType", None),
            field(item, "System.State", None),
        ))


def field(wi, name, default=""):
    val = wi.get("fields", {}).get(name, default)
    if isinstance(val, dict):  # e.g. AssignedTo is a dict with displayName
        return val.get("displayName", default)
    return val if val is not None else default


def write_to_workbook(items):
    wb = openpyxl.load_workbook(WORKBOOK_PATH)
    if "Raw Data" not in wb.sheetnames:
        sys.exit('ERROR: "Raw Data" sheet not found in the workbook. '
                  "Use the workbook Claude generated, or create that tab first.")
    ws = wb["Raw Data"]

    header_row = 5
    data_start = header_row + 1
    columns = [
        ("Work Item ID", "System.Id"),
        ("Title", "System.Title"),
        ("Work Item Type", "System.WorkItemType"),
        ("State", "System.State"),
        ("State Category", "_state_category"),
        ("Board Column", "System.BoardColumn"),
        ("Board Column Done", "System.BoardColumnDone"),
        ("Board Lane", "System.BoardLane"),
        ("Assigned To", "System.AssignedTo"),
        ("Iteration Path", "System.IterationPath"),
        ("Area Path", "System.AreaPath"),
        ("Story Points", "Microsoft.VSTS.Scheduling.StoryPoints"),
        ("Priority", "Microsoft.VSTS.Common.Priority"),
        ("Created Date", "System.CreatedDate"),
        ("Changed Date", "System.ChangedDate"),
        ("Tags", "System.Tags"),
        ("URL", None),
        ("Parent ID", "System.Parent"),
        ("State Change Date", "Microsoft.VSTS.Common.StateChangeDate"),
        ("Closed Date", "Microsoft.VSTS.Common.ClosedDate"),
    ]

    # Rebuild the source columns explicitly, preventing stale values from a
    # previous wider or narrower pull.
    clear_to_col = max(ws.max_column, len(columns))
    clear_to_row = max(ws.max_row, data_start + len(items) + 5)
    for row in range(header_row, clear_to_row + 1):
        for col in range(1, clear_to_col + 1):
            ws.cell(row=row, column=col).value = None
    for col, (label, _) in enumerate(columns, start=1):
        ws.cell(row=header_row, column=col, value=label)

    org_project_url = f"https://dev.azure.com/{ORG_NAME}/{PROJECT_NAME}/_workitems/edit/"
    date_fields = {
        "System.CreatedDate", "System.ChangedDate",
        "Microsoft.VSTS.Common.StateChangeDate",
        "Microsoft.VSTS.Common.ClosedDate",
    }

    for offset, wi in enumerate(items):
        row = data_start + offset
        wi_id = wi.get("id")
        for col, (_, azure_field) in enumerate(columns, start=1):
            if azure_field is None:
                value = f"{org_project_url}{wi_id}" if wi_id else ""
            elif azure_field == "System.Id":
                value = wi_id
            elif azure_field == "_state_category":
                value = wi.get("_state_category")
            else:
                value = field(wi, azure_field, None)
                if azure_field in date_fields and value:
                    value = value[:10]
            ws.cell(row=row, column=col, value=value)

    # Stamp when this was last pulled
    ws["A2"] = (
        f"Last pulled from Azure DevOps: "
        f"{datetime.datetime.now().strftime('%d/%m/%Y %H:%M')}. "
        "Do not edit by hand — this tab is overwritten on each run."
    )

    wb.save(WORKBOOK_PATH)


def main():
    print(f"Querying work items for project '{PROJECT_NAME}'"
          + (f" (iteration: {ITERATION_PATH})" if ITERATION_PATH else " (all)") + "...")
    ids = run_wiql()
    print(f"Found {len(ids)} work items.")
    if not ids:
        print("Nothing to write. Check ORG_NAME / PROJECT_NAME / ITERATION_PATH.")
        return
    items = get_work_items_batch(ids)
    enrich_state_categories(items)
    write_to_workbook(items)
    print(f"Wrote {len(items)} work items into '{WORKBOOK_PATH}' -> Raw Data tab.")

    # Keep the workbook as a one-command Azure mirror: every successful pull
    # immediately rebuilds all analytical sheets from the fresh Raw Data.
    dashboard_script = os.path.join(os.path.dirname(__file__), "build_dashboard.py")
    if os.path.exists(dashboard_script):
        import subprocess
        subprocess.run([sys.executable, dashboard_script], check=True)
        print("Rebuilt all delivery dashboards from the fresh Azure data.")
    else:
        print("WARNING: build_dashboard.py not found; Raw Data was updated only.")

    print("Open the workbook in Excel — calculations and charts are ready.")


if __name__ == "__main__":
    main()