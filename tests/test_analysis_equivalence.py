"""
Phase 1 extraction verification.

Proves core/analysis.py produces output IDENTICAL to the original inline
functions that lived in dashboard_app.py before the Phase 1 split (see
`git show <pre-refactor-commit>:dashboard_app.py`). Golden reference
implementations below are copied verbatim (byte-for-byte body) from that
original file so this test is independent of any bug that might be
(re)introduced into core/analysis.py later.

Fixture data: the cached Delivery_Manager_Dashboard.xlsx workbook's "Raw
Data" sheet (used as the AZDO_PAT-unset fallback), loaded through the same
row-parsing logic dashboard_app.read_workbook_items() uses.
"""

import datetime as dt
import os
from collections import Counter, defaultdict

import openpyxl
import pandas as pd
import pytest

import dashboard_theme
from core import analysis

WORKBOOK = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                         "Delivery_Manager_Dashboard.xlsx")

pytestmark = pytest.mark.skipif(
    not os.path.exists(WORKBOOK),
    reason="Delivery_Manager_Dashboard.xlsx fixture not present in this checkout",
)


# --------------------------------------------------------------------------
# Fixture loading: mirrors dashboard_app.read_workbook_items() row parsing
# exactly (that function's logic did not move — only the _category/_sprint/
# _leaf/_parse_date helpers it calls moved to core.analysis).
# --------------------------------------------------------------------------
def _load_workbook_items():
    wb = openpyxl.load_workbook(WORKBOOK, data_only=True)
    ws = wb["Raw Data"]
    header = [c.value for c in ws[5]]
    idx = {name: i for i, name in enumerate(header) if name}
    items = []
    for row in ws.iter_rows(min_row=6, values_only=True):
        if not row or row[0] is None and (row[1] is None):
            continue

        def g(name, default=None, _row=row):
            i = idx.get(name)
            return _row[i] if i is not None and i < len(_row) else default

        created = analysis._parse_date(g("Created Date"))
        state = g("State") or "Unknown"
        items.append({
            "id": g("Work Item ID"),
            "title": g("Title") or "",
            "type": g("Work Item Type") or "Unknown",
            "state": state,
            "state_category": analysis._category(state, g("State Category")),
            "board_column": g("Board Column") or state,
            "board_column_done": bool(g("Board Column Done")),
            "board_lane": g("Board Lane") or "Default",
            "assignee": g("Assigned To") or "Unassigned",
            "sprint": analysis._sprint(g("Iteration Path")),
            "area": analysis._leaf(g("Area Path") or analysis.PROJECT),
            "sp": g("Story Points"),
            "priority": g("Priority"),
            "created": created,
            "changed": analysis._parse_date(g("Changed Date")),
            "parent": g("Parent ID"),
            "tags": [t.strip() for t in str(g("Tags") or "").split(";") if t.strip()],
            "url": f"https://dev.azure.com/{analysis.ORG}/{analysis.PROJECT}/_workitems/edit/{g('Work Item ID')}",
        })
    return items


@pytest.fixture(scope="module")
def all_items():
    items = _load_workbook_items()
    assert items, "workbook fixture produced zero items — check the workbook path/sheet"
    return items


@pytest.fixture(scope="module")
def dev_items(all_items):
    return [i for i in all_items if i["type"] in analysis.DEV_TYPES]


# ==========================================================================
# GOLDEN reference implementations — copied verbatim from the pre-Phase-1
# dashboard_app.py (git show 6941bbc:dashboard_app.py), constants inlined.
# ==========================================================================
_STALE_DAYS = 14
_TERMINAL_CATEGORIES = {"Completed", "Removed"}
_DELIVERY_TYPES = ("Epic", "Feature", "User Story", "Task", "Bug",
                    "Test Case", "Test Suite", "Test Plan")


def golden_is_done(i):
    return i.get("state_category") == "Completed"


def golden_is_open(i):
    return i.get("state_category") not in _TERMINAL_CATEGORIES


def golden_is_active(i):
    return i.get("state_category") == "InProgress"


def golden_item_metrics(items):
    total = len(items)
    done = sum(golden_is_done(i) for i in items)
    active = sum(golden_is_active(i) for i in items)
    return {
        "total": total,
        "done": done,
        "active": active,
        "open": sum(golden_is_open(i) for i in items),
        "unassigned": sum(
            i["type"] == "Task" and i["assignee"] == "Unassigned"
            for i in items
        ),
        "stale": sum(golden_is_open(i) and i["created"] and (dt.date.today() - i["created"]).days >= _STALE_DAYS for i in items),
        "done_pct": done / total if total else 0,
    }


def golden_type_progress(items):
    children = defaultdict(list)
    for i in items:
        if i["parent"] is not None:
            children[i["parent"]].append(i)
    hier = any(i["parent"] is not None for i in items)

    memo = {}
    visiting = set()

    def effective(item):
        item_id = item["id"]
        if item_id in memo:
            return memo[item_id]
        if item_id in visiting:
            return False
        visiting.add(item_id)
        child_items = children.get(item_id, [])
        result = (
            all(effective(child) for child in child_items)
            if child_items
            else golden_is_done(item)
        )
        visiting.remove(item_id)
        memo[item_id] = result
        return result

    agg = defaultdict(lambda: {"total": 0, "done": 0})
    for t in _DELIVERY_TYPES:
        agg[t] = {"total": 0, "done": 0}
    for i in items:
        d = effective(i) if hier else golden_is_done(i)
        agg[i["type"]]["total"] += 1
        agg[i["type"]]["done"] += 1 if d else 0
    out = {}
    for t, m in agg.items():
        out[t] = {"total": m["total"], "done": m["done"],
                  "pct": m["done"] / m["total"] if m["total"] else None}
    out["hierarchy_used"] = hier
    return out


def golden_scope_metrics(items):
    stories = [i for i in items if i["type"] == "User Story"]
    tasks = [i for i in items if i["type"] == "Task"]
    tasks_by_parent = defaultdict(list)
    hier = any(i["parent"] is not None for i in items)
    for t in tasks:
        if t["parent"] is not None:
            tasks_by_parent[t["parent"]].append(t)
    done_story_items = []
    for story in stories:
        child_tasks = tasks_by_parent.get(story["id"])
        story_done = (
            all(golden_is_done(task) for task in child_tasks)
            if child_tasks and hier
            else golden_is_done(story)
        )
        if story_done:
            done_story_items.append(story)

    task_done_count = sum(golden_is_done(task) for task in tasks)
    total_sp = sum(float(story["sp"] or 0) for story in stories)
    done_sp = sum(float(story["sp"] or 0) for story in done_story_items)
    return {
        "stories": len(stories),
        "stories_done": len(done_story_items),
        "scope_pct": len(done_story_items) / len(stories) if stories else None,
        "tasks": len(tasks),
        "tasks_done": task_done_count,
        "task_pct": task_done_count / len(tasks) if tasks else None,
        "total_sp": total_sp,
        "done_sp": done_sp,
        "velocity_pct": done_sp / total_sp if total_sp else None,
        "hier": hier,
    }


def golden_percent(value):
    return round(value * 100, 1) if value is not None else None


def golden_ribbon(scope, task, unassigned, stale, accents):
    red = 0
    if (scope or 0) < 0.4:
        red += 1
    if (task or 0) < 0.4:
        red += 1
    if unassigned and unassigned >= 25:
        red += 1
    if stale and stale >= 10:
        red += 1
    if red >= 3:
        return "CRITICAL", accents["red"]
    if red >= 2:
        return "AT RISK", accents["amber"]
    return "HEALTHY", accents["green"]


def golden_weekly_creation_closure(delivery_items):
    created = Counter()
    closed = Counter()
    for item in delivery_items:
        if not item.get("created"):
            continue
        monday = item["created"] - dt.timedelta(days=item["created"].weekday())
        created[monday] += 1
        closure = item.get("changed") if golden_is_done(item) else None
        if closure:
            closed_monday = closure - dt.timedelta(days=closure.weekday())
            closed[closed_monday] += 1
    weeks = sorted(set(created) | set(closed))
    return pd.DataFrame([
        {
            "period": week.strftime("%d %b"),
            "Created": created.get(week, 0),
            "Closed": closed.get(week, 0),
        }
        for week in weeks[-12:]
    ])


def golden_sprint_summary_df(dev):
    groups = defaultdict(list)
    for i in dev:
        groups[i["sprint"]].append(i)
    rows = []
    for sprint in sorted(groups, key=lambda s: (s == analysis.PB, s)):
        m = groups[sprint]
        scg = golden_scope_metrics(m)
        im = golden_item_metrics(m)
        rows.append({
            "Iteration": sprint,
            "Total Dev Items": len(m),
            "User Stories": scg["stories"],
            "Stories Done": scg["stories_done"],
            "Scope Done %": golden_percent(scg["scope_pct"]),
            "Tasks": scg["tasks"],
            "Tasks Done": scg["tasks_done"],
            "Task Done %": golden_percent(scg["task_pct"]),
            "Active": im["active"],
            "Unassigned": im["unassigned"],
            f"Open ≥{_STALE_DAYS}d": im["stale"],
            "Iteration Meaning": "No sprint assigned" if sprint == analysis.PB else "Committed iteration",
        })
    return pd.DataFrame(rows)


def golden_team_df(dev):
    tasks = [item for item in dev if item["type"] == "Task"]
    tasks_by_assignee = defaultdict(list)
    tasks_by_story = defaultdict(list)
    for task in tasks:
        tasks_by_assignee[task["assignee"]].append(task)
        if task["parent"] is not None:
            tasks_by_story[task["parent"]].append(task)

    completed_story_ids = {
        story_id
        for story_id, child_tasks in tasks_by_story.items()
        if child_tasks and all(golden_is_done(task) for task in child_tasks)
    }

    rows = []
    ordered_groups = sorted(
        tasks_by_assignee.items(), key=lambda group: (-len(group[1]), group[0])
    )
    for assignee, owned_tasks in ordered_groups:
        done_count = sum(golden_is_done(task) for task in owned_tasks)
        involved_story_ids = {
            task["parent"] for task in owned_tasks if task["parent"] is not None
        }
        rows.append({
            "Assignee": assignee,
            "Tasks": len(owned_tasks),
            "Tasks Done": done_count,
            "Task Completion %": golden_percent(done_count / len(owned_tasks)),
            "Active": sum(golden_is_active(task) for task in owned_tasks),
            "Open": sum(golden_is_open(task) for task in owned_tasks),
            f"Open ≥{_STALE_DAYS}d": golden_item_metrics(owned_tasks)["stale"],
            "Stories Involved": len(involved_story_ids),
            "Stories Fully Done": len(involved_story_ids & completed_story_ids),
            "Areas": ", ".join(sorted({task["area"] for task in owned_tasks})),
        })
    return pd.DataFrame(rows)


def golden_area_df(dev):
    rows = []
    for area in sorted({i["area"] for i in dev}, key=lambda a: -sum(1 for i in dev if i["area"] == a)):
        members = [i for i in dev if i["area"] == area]
        sc = golden_scope_metrics(members)
        im = golden_item_metrics(members)
        rows.append({
            "Area": area, "Total": len(members),
            "Stories": sc["stories"], "Stories Done": sc["stories_done"],
            "Scope %": golden_percent(sc["scope_pct"]),
            "Tasks": sc["tasks"], "Tasks Done": sc["tasks_done"],
            "Task %": golden_percent(sc["task_pct"]),
            "SP": sc["total_sp"], "Done SP": sc["done_sp"],
            "Active": im["active"], "Unassigned": im["unassigned"],
            f"Open ≥{_STALE_DAYS}d": im["stale"],
        })
    return pd.DataFrame(rows)


# ==========================================================================
# Equivalence assertions
# ==========================================================================
def test_business_rules_constants_unchanged():
    assert analysis.STALE_DAYS == 14
    assert analysis.PB == "Product Backlog"
    assert analysis.DELIVERY_TYPES == _DELIVERY_TYPES
    assert analysis.TERMINAL_CATEGORIES == _TERMINAL_CATEGORIES


def test_is_done_is_open_is_active_identical(dev_items):
    for i in dev_items:
        assert analysis.is_done(i) == golden_is_done(i)
        assert analysis.is_open(i) == golden_is_open(i)
        assert analysis.is_active(i) == golden_is_active(i)


def test_item_metrics_identical(dev_items):
    assert analysis.item_metrics(dev_items) == golden_item_metrics(dev_items)


def test_item_metrics_identical_per_sprint_subset(dev_items):
    # Exercise more code paths (empty/small groups) the way sprint/tag/area
    # views do, not just the full delivery set.
    groups = defaultdict(list)
    for i in dev_items:
        groups[i["sprint"]].append(i)
    for members in groups.values():
        assert analysis.item_metrics(members) == golden_item_metrics(members)


def test_type_progress_identical(dev_items):
    assert analysis.type_progress(dev_items) == golden_type_progress(dev_items)


def test_scope_metrics_identical(dev_items):
    assert analysis.scope_metrics(dev_items) == golden_scope_metrics(dev_items)


def test_scope_metrics_identical_per_sprint_subset(dev_items):
    groups = defaultdict(list)
    for i in dev_items:
        groups[i["sprint"]].append(i)
    for members in groups.values():
        assert analysis.scope_metrics(members) == golden_scope_metrics(members)


def test_scope_metrics_identical_per_area_subset(dev_items):
    areas = sorted({i["area"] for i in dev_items})
    for area in areas:
        members = [i for i in dev_items if i["area"] == area]
        assert analysis.scope_metrics(members) == golden_scope_metrics(members)


def test_percent_identical():
    for value in (None, 0, 0.333333, 1, 0.4051, 0.995):
        assert analysis.percent(value) == golden_percent(value)


def test_ribbon_identical_across_grid():
    accents = dashboard_theme.ACCENTS
    scope_values = (None, 0, 0.1, 0.39, 0.4, 0.6, 1)
    task_values = (None, 0, 0.1, 0.39, 0.4, 0.6, 1)
    unassigned_values = (0, 1, 24, 25, 30)
    stale_values = (0, 1, 9, 10, 20)
    for scope in scope_values:
        for task in task_values:
            for unassigned in unassigned_values:
                for stale in stale_values:
                    verdict, color = analysis.ribbon(scope, task, unassigned, stale)
                    expected_verdict, expected_color = golden_ribbon(
                        scope, task, unassigned, stale, accents
                    )
                    assert verdict == expected_verdict
                    assert color == expected_color


def test_weekly_creation_closure_identical(dev_items):
    actual = analysis.weekly_creation_closure(dev_items)
    expected = golden_weekly_creation_closure(dev_items)
    pd.testing.assert_frame_equal(
        actual.reset_index(drop=True), expected.reset_index(drop=True)
    )


def test_sprint_summary_df_identical(dev_items):
    actual = analysis.sprint_summary_df(dev_items)
    expected = golden_sprint_summary_df(dev_items)
    pd.testing.assert_frame_equal(
        actual.reset_index(drop=True), expected.reset_index(drop=True)
    )


def test_team_df_identical(dev_items):
    actual = analysis.team_df(dev_items)
    expected = golden_team_df(dev_items)
    pd.testing.assert_frame_equal(
        actual.reset_index(drop=True), expected.reset_index(drop=True)
    )


def test_area_df_identical(dev_items):
    actual = analysis.area_df(dev_items)
    expected = golden_area_df(dev_items)
    pd.testing.assert_frame_equal(
        actual.reset_index(drop=True), expected.reset_index(drop=True)
    )


def test_normalization_helpers_identical(all_items):
    # _leaf / _sprint / _parse_date / _category / _wi are exercised via the
    # fixture loader already (identical inputs -> identical dict shape);
    # spot check the pure helpers directly too.
    assert analysis._leaf("Hoteliana\\Supplier Dashboard") == "Supplier Dashboard"
    assert analysis._leaf("") == ""
    assert analysis._sprint("Hoteliana\\Sprint 1") == "Sprint 1"
    assert analysis._sprint("") == analysis.PB
    assert analysis._parse_date("2026-08-13") == dt.date(2026, 8, 13)
    assert analysis._parse_date(None) is None
    assert analysis._category("Closed", None) == "Completed"
    assert analysis._category("Active", None) == "InProgress"
    assert analysis._category("SomeCustomState", "Proposed") == "Proposed"


def test_hierarchy_rollup_story_done_only_when_all_tasks_done(dev_items):
    """Business rule: 'A Story is Done only when all linked Tasks are Done.'"""
    stories = [i for i in dev_items if i["type"] == "User Story"]
    tasks = [i for i in dev_items if i["type"] == "Task"]
    tasks_by_parent = defaultdict(list)
    for t in tasks:
        if t["parent"] is not None:
            tasks_by_parent[t["parent"]].append(t)

    scope = analysis.scope_metrics(dev_items)
    hier = any(i["parent"] is not None for i in dev_items)
    if not hier:
        pytest.skip("fixture has no Parent ID links; hierarchy roll-up not exercised")

    expected_done_ids = set()
    for story in stories:
        children = tasks_by_parent.get(story["id"])
        done = (
            all(analysis.is_done(c) for c in children)
            if children
            else analysis.is_done(story)
        )
        if done:
            expected_done_ids.add(story["id"])

    assert scope["stories_done"] == len(expected_done_ids)


def test_unassigned_risk_counts_only_tasks_not_stories(dev_items):
    """Business rule: 'A Task is the unassigned risk, not a Story.'"""
    metrics = analysis.item_metrics(dev_items)
    manual_unassigned_tasks = sum(
        1 for i in dev_items if i["type"] == "Task" and i["assignee"] == "Unassigned"
    )
    manual_unassigned_stories = sum(
        1 for i in dev_items if i["type"] == "User Story" and i["assignee"] == "Unassigned"
    )
    assert metrics["unassigned"] == manual_unassigned_tasks
    # An unassigned User Story must never be counted in this metric, even if
    # some exist in the fixture.
    if manual_unassigned_stories:
        assert metrics["unassigned"] != manual_unassigned_tasks + manual_unassigned_stories
