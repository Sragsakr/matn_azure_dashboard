"""
Phase 4 verification: core.analysis.apply_global_filters().

Tests the pure global-filter predicate used by components/filter_bar.py
and the Team Analysis click-to-filter grid. Follows the fixture pattern in
tests/test_analysis_equivalence.py — real workbook data for realistic
combination coverage, plus small synthetic item lists for exact edge-case
control (empty filters, date boundaries, missing `created`).
"""

import datetime as dt
import os

import openpyxl
import pytest

from core import analysis

WORKBOOK = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                         "Delivery_Manager_Dashboard.xlsx")

pytestmark = pytest.mark.skipif(
    not os.path.exists(WORKBOOK),
    reason="Delivery_Manager_Dashboard.xlsx fixture not present in this checkout",
)


def _load_workbook_items():
    wb = openpyxl.load_workbook(WORKBOOK, data_only=True)
    ws = wb["Raw Data"]
    header = [c.value for c in ws[5]]
    idx = {name: i for i, name in enumerate(header) if name}
    items = []
    for row in ws.iter_rows(min_row=6, values_only=True):
        if not row or row[0] is None and (row[1] is None):
            continue

        def g(name, default=None, _row=row):
            i = idx.get(name)
            return _row[i] if i is not None and i < len(_row) else default

        created = analysis._parse_date(g("Created Date"))
        state = g("State") or "Unknown"
        items.append({
            "id": g("Work Item ID"),
            "title": g("Title") or "",
            "type": g("Work Item Type") or "Unknown",
            "state": state,
            "state_category": analysis._category(state, g("State Category")),
            "board_column": g("Board Column") or state,
            "board_column_done": bool(g("Board Column Done")),
            "board_lane": g("Board Lane") or "Default",
            "assignee": g("Assigned To") or "Unassigned",
            "sprint": analysis._sprint(g("Iteration Path")),
            "area": analysis._leaf(g("Area Path") or analysis.PROJECT),
            "sp": g("Story Points"),
            "priority": g("Priority"),
            "created": created,
            "changed": analysis._parse_date(g("Changed Date")),
            "parent": g("Parent ID"),
            "tags": [t.strip() for t in str(g("Tags") or "").split(";") if t.strip()],
            "url": f"https://dev.azure.com/{analysis.ORG}/{analysis.PROJECT}/_workitems/edit/{g('Work Item ID')}",
        })
    return items


@pytest.fixture(scope="module")
def dev_items():
    items = _load_workbook_items()
    dev = [i for i in items if i["type"] in analysis.DEV_TYPES]
    assert dev, "workbook fixture produced zero dev items — check the workbook path/sheet"
    return dev


# ==========================================================================
# Real-fixture combination coverage
# ==========================================================================
def test_no_filters_returns_all_items(dev_items):
    assert analysis.apply_global_filters(dev_items, {}) == dev_items
    assert analysis.apply_global_filters(dev_items, {
        "sprints": [], "assignees": [], "types": [],
        "date_from": None, "date_to": None,
    }) == dev_items


def test_filter_by_sprint_only(dev_items):
    target_sprint = dev_items[0]["sprint"]
    result = analysis.apply_global_filters(dev_items, {"sprints": [target_sprint]})
    assert result, "expected at least one match for a sprint present in the fixture"
    assert all(i["sprint"] == target_sprint for i in result)
    expected = [i for i in dev_items if i["sprint"] == target_sprint]
    assert result == expected


def test_filter_by_assignee_only(dev_items):
    target_assignee = next(i["assignee"] for i in dev_items if i["assignee"])
    result = analysis.apply_global_filters(dev_items, {"assignees": [target_assignee]})
    assert result
    assert all(i["assignee"] == target_assignee for i in result)
    expected = [i for i in dev_items if i["assignee"] == target_assignee]
    assert result == expected


def test_filter_by_type_only(dev_items):
    target_type = dev_items[0]["type"]
    result = analysis.apply_global_filters(dev_items, {"types": [target_type]})
    assert result
    assert all(i["type"] == target_type for i in result)
    expected = [i for i in dev_items if i["type"] == target_type]
    assert result == expected


def test_filter_by_multiple_values_in_one_facet_is_or(dev_items):
    types_present = sorted({i["type"] for i in dev_items})
    if len(types_present) < 2:
        pytest.skip("fixture has fewer than 2 distinct types")
    picked = types_present[:2]
    result = analysis.apply_global_filters(dev_items, {"types": picked})
    assert all(i["type"] in picked for i in result)
    expected = [i for i in dev_items if i["type"] in picked]
    assert result == expected


def test_filter_by_date_range(dev_items):
    dated = [i for i in dev_items if i["created"]]
    if not dated:
        pytest.skip("fixture has no dated items")
    dates = sorted(i["created"] for i in dated)
    mid = dates[len(dates) // 2]
    result = analysis.apply_global_filters(dev_items, {"date_from": mid, "date_to": dates[-1]})
    assert all(i["created"] and mid <= i["created"] <= dates[-1] for i in result)
    expected = [i for i in dev_items if i["created"] and mid <= i["created"] <= dates[-1]]
    assert result == expected


def test_combined_filters_are_and_across_facets(dev_items):
    target_type = dev_items[0]["type"]
    same_type_items = [i for i in dev_items if i["type"] == target_type]
    target_sprint = same_type_items[0]["sprint"]
    result = analysis.apply_global_filters(
        dev_items, {"types": [target_type], "sprints": [target_sprint]}
    )
    expected = [
        i for i in dev_items if i["type"] == target_type and i["sprint"] == target_sprint
    ]
    assert result == expected
    assert all(i["type"] == target_type and i["sprint"] == target_sprint for i in result)


def test_filter_matching_nothing_returns_empty_list(dev_items):
    result = analysis.apply_global_filters(dev_items, {"assignees": ["__no_such_person__"]})
    assert result == []


# ==========================================================================
# Synthetic edge cases (exact control over `created`/missing fields)
# ==========================================================================
def _item(**overrides):
    base = {
        "id": 1, "title": "t", "type": "Task", "state": "Active",
        "state_category": "InProgress", "assignee": "Alice", "sprint": "Sprint 1",
        "area": "A", "sp": None, "priority": None, "created": dt.date(2026, 1, 10),
        "changed": None, "parent": None, "tags": [],
    }
    base.update(overrides)
    return base


def test_date_from_only_is_inclusive_lower_bound():
    items = [
        _item(id=1, created=dt.date(2026, 1, 9)),
        _item(id=2, created=dt.date(2026, 1, 10)),
        _item(id=3, created=dt.date(2026, 1, 11)),
    ]
    result = analysis.apply_global_filters(items, {"date_from": dt.date(2026, 1, 10)})
    assert [i["id"] for i in result] == [2, 3]


def test_date_to_only_is_inclusive_upper_bound():
    items = [
        _item(id=1, created=dt.date(2026, 1, 9)),
        _item(id=2, created=dt.date(2026, 1, 10)),
        _item(id=3, created=dt.date(2026, 1, 11)),
    ]
    result = analysis.apply_global_filters(items, {"date_to": dt.date(2026, 1, 10)})
    assert [i["id"] for i in result] == [1, 2]


def test_items_with_no_created_date_excluded_when_date_filter_active():
    items = [
        _item(id=1, created=dt.date(2026, 1, 10)),
        _item(id=2, created=None),
    ]
    result = analysis.apply_global_filters(items, {"date_from": dt.date(2026, 1, 1)})
    assert [i["id"] for i in result] == [1]


def test_items_with_no_created_date_kept_when_no_date_filter():
    items = [
        _item(id=1, created=dt.date(2026, 1, 10)),
        _item(id=2, created=None),
    ]
    result = analysis.apply_global_filters(items, {"assignees": ["Alice"]})
    assert [i["id"] for i in result] == [1, 2]


def test_empty_items_list_returns_empty_list():
    assert analysis.apply_global_filters([], {"sprints": ["Sprint 1"]}) == []
    assert analysis.apply_global_filters([], {}) == []


def test_apply_global_filters_does_not_mutate_input():
    items = [_item(id=1), _item(id=2, assignee="Bob")]
    original = list(items)
    analysis.apply_global_filters(items, {"assignees": ["Alice"]})
    assert items == original
